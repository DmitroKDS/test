import flet
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import fnmatch
import math
import pymysql
import os
pymysql.install_as_MySQLdb()

def OpenDB():
    global DBConnector, DBCursor
    DBConnector = pymysql.connect(
        host='mf5f.mysql.ukraine.com.ua',
        user='mf5f_gloves0stock',
        password=os.getenv("MY_VAR"),
        database='mf5f_gloves0stock',
        port=3306,
        charset='utf8mb4',
        cursorclass=pymysql.cursors.Cursor  # або DictCursor, якщо потрібні dict-и
    )
    DBCursor = DBConnector.cursor()

def CloseDB():
    DBCursor.close()
    DBConnector.close()

CurrentPageStatus ={"/workers": None, "/worker_shifts":None, "/worker_shift_gloves":None, "/worker_shift_machine_gloves":None, "/products_for_machines":None}
PageRequests = {"/workers": {"Normal":"SELECT Id, Name, Stage, Password, Exist FROM workers ORDER BY Exist DESC, Id DESC", "Sort":"SELECT Id, Name, Stage, Password, Exist FROM workers ORDER BY Exist DESC, (SortColumn) (SortDirection)", "Filter":"SELECT Id, Name, Stage, Password, Exist FROM workers WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, Id DESC", "WithDeleted":"SELECT Id, Name, Stage, Password, Exist FROM workers WHERE Exist=1 ORDER BY Exist DESC, Id DESC", "SortFilter":"SELECT Id, Name, Stage, Password, Exist FROM workers WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, (SortColumn) (SortDirection)", "SortWithDeleted":"SELECT Id, Name, Stage, Password, Exist FROM workers WHERE Exist=1 ORDER BY Exist DESC, (SortColumn) (SortDirection)", "FilterWithDeleted":"SELECT Id, Name, Stage, Password, Exist FROM workers WHERE Exist=1 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, Id DESC", "SortFilterWithDeleted":"SELECT Id, Name, Stage, Password, Exist FROM workers WHERE Exist=1 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, (SortColumn) (SortDirection)"}, 
                "/worker_shifts": {"Normal":"SELECT Id, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments FROM workers_shifts WHERE WorkerId = (ActiveId) ORDER BY Id DESC", "Sort":"SELECT Id, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments FROM workers_shifts WHERE WorkerId = (ActiveId) ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Id, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments FROM workers_shifts WHERE WorkerId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Id DESC", "SortFilter":"SELECT Id, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments FROM workers_shifts WHERE WorkerId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/worker_shift_gloves": {"Normal":"SELECT Id, Machine, ProductId, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) ORDER BY Id DESC", "Sort":"SELECT Id, Machine, ProductId, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Id, Machine, ProductId, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Id DESC", "SortFilter":"SELECT Id, Machine, ProductId, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/worker_shift_machine_gloves": {"Normal":"SELECT Machine, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) ORDER BY Id DESC", "Sort":"SELECT Machine, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Machine, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Id DESC", "SortFilter":"SELECT Machine, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE WorkerId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/products_for_machines": {"Normal":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines ORDER BY Exist DESC, Id DESC", "Sort":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines ORDER BY Exist DESC, (SortColumn) (SortDirection)", "Filter":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, Id DESC", "WithDeleted":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines WHERE Exist=1 ORDER BY Exist DESC, Id DESC", "SortFilter":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, (SortColumn) (SortDirection)", "SortWithDeleted":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines WHERE Exist=1 ORDER BY Exist DESC, (SortColumn) (SortDirection)", "FilterWithDeleted":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines WHERE Exist=1 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, Id DESC", "SortFilterWithDeleted":"SELECT Id, Machine, ProductId, TimeStart, TimeEnd, Exist FROM products_for_machines WHERE Exist=1 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, (SortColumn) (SortDirection)"}, 
                "/stage_machines_gloves":{"Normal":"SELECT Machine, ProductId, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE Stage = '(ActiveStage)' ORDER BY Id DESC", "Sort":"SELECT Machine, Product, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE Stage = '(ActiveStage)' ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Machine, Product, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE Stage = '(ActiveStage)' AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Id DESC", "SortFilter":"SELECT Machine, Product, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE Stage = '(ActiveStage)' AND (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/products": {"Normal":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products ORDER BY Exist DESC, Id DESC", "Sort":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products ORDER BY Exist DESC, (SortColumn) (SortDirection)", "Filter":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, Id DESC", "WithDeleted":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products WHERE Exist=1 ORDER BY Exist DESC, Id DESC", "SortFilter":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, (SortColumn) (SortDirection)", "SortWithDeleted":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products WHERE Exist=1 ORDER BY Exist DESC, (SortColumn) (SortDirection)", "FilterWithDeleted":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products WHERE Exist=1 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, Id DESC", "SortFilterWithDeleted":"SELECT Id, Artikel, FullName, ShortName, Come, Stage, Exist FROM products WHERE Exist=1 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Exist DESC, (SortColumn) (SortDirection)"},
                "/product_info": {"Normal":"SELECT Machine, Stage, Sort, Pair, AddDate FROM products_gloves_quantity WHERE ProductId = (ActiveId) ORDER BY Id DESC", "Sort":"SELECT Machine, Stage, Sort, Pair, AddDate FROM products_gloves_quantity WHERE ProductId = (ActiveId) ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Machine, Stage, Sort, Pair, AddDate FROM products_gloves_quantity WHERE ProductId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Id DESC", "SortFilter":"SELECT Machine, Stage, Sort, Pair, AddDate FROM products_gloves_quantity WHERE ProductId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/comings": {"Normal":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings ORDER BY TimeEnd DESC, Id DESC", "Sort":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings ORDER BY TimeEnd DESC, (SortColumn) (SortDirection)", "Filter":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY TimeEnd DESC, Id DESC", "WithDeleted":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings WHERE Pair!=0 ORDER BY TimeEnd DESC, Id DESC", "SortFilter":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY TimeEnd DESC, (SortColumn) (SortDirection)", "SortWithDeleted":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings WHERE Pair!=0 ORDER BY TimeEnd DESC, (SortColumn) (SortDirection)", "FilterWithDeleted":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings WHERE Pair!=0 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY TimeEnd DESC, Id DESC", "SortFilterWithDeleted":"SELECT Id, Stage, Product, Pair, TimeStart, TimeEnd FROM comings WHERE Pair!=0 AND (FilterColumn) LIKE '(FilterValue)' ORDER BY TimeEnd DESC, (SortColumn) (SortDirection)"}, 
                "/coming_info": {"Normal":"SELECT Stage, WorkerName, Product, Pair, Sort, AddDate FROM comings_info WHERE ComeId = (ActiveId) ORDER BY Id DESC", "Sort":"SELECT Stage, WorkerName, Product, Pair, Sort, AddDate FROM comings_info WHERE ComeId = (ActiveId) ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Stage, WorkerName, Product, Pair, Sort, AddDate FROM comings_info WHERE ComeId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY Id DESC", "SortFilter":"SELECT Stage, WorkerName, Product, Pair, Sort, AddDate FROM comings_info WHERE ComeId = (ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/schedule": {"Normal":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule ORDER BY Id DESC", "Sort":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule", "Filter":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule ORDER BY Id DESC", "SortFilter":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule"}, 
                "/downloaded_schedule": {"Normal":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule ORDER BY Id DESC", "Sort":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY Id DESC", "SortFilter":"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/salary": {"Normal":"SELECT ShiftStart FROM workers_shifts WHERE Id = (MinShiftId)", "Sort":"SELECT ShiftStart FROM workers_shifts WHERE Id = (MinShiftId)", "Filter":f"SELECT ShiftStart FROM workers_shifts WHERE Id = (MinShiftId)", "SortFilter":"SELECT ShiftStart FROM workers_shifts WHERE Id = (MinShiftId)"},
                "/shifts_salary":{"Normal":"SELECT ShiftId, WorkerId, ShiftPlan, PairPrice FROM salary", "Sort":"SELECT ShiftId, WorkerId, ShiftPlan, PairPrice FROM salary", "Filter":"SELECT ShiftId, WorkerId, ShiftPlan, PairPrice FROM salary", "SortFilter":"SELECT ShiftId, WorkerId, ShiftPlan, PairPrice FROM salary"},
                "/salary_set":{"Normal":"SELECT Id, Name, Stage FROM workers WHERE Exist=1 ORDER BY Id DESC"},
                "/unloadings":{"Normal":"SELECT Id, Name, Stage, Date FROM unloadings", "Sort":"SELECT Id, Name, Stage, Date FROM unloadings ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Id, Name, Stage, Date FROM unloadings WHERE (FilterColumn) LIKE '(FilterValue)'", "SortFilter":"SELECT Id, Name, Stage, Date FROM unloadings WHERE (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},
                "/unloading_info":{"Normal":"SELECT Id, ProductId, Pair FROM unloadings_info WHERE UnloadingId=(ActiveId)", "Sort":"SELECT Id, ProductId, Pair FROM unloadings_info WHERE UnloadingId=(ActiveId) ORDER BY (SortColumn) (SortDirection)", "Filter":"SELECT Id, ProductId, Pair FROM unloadings_info WHERE UnloadingId=(ActiveId) AND (FilterColumn) LIKE '(FilterValue)'", "SortFilter":"SELECT Id, ProductId, Pair FROM unloadings_info WHERE UnloadingId=(ActiveId) AND (FilterColumn) LIKE '(FilterValue)' ORDER BY (SortColumn) (SortDirection)"},}

DeleteRequests = {"/workers":[f"UPDATE workers SET Exist = 0 WHERE Id = (DeleteId)", f"SELECT Name, Stage FROM workers WHERE Id = (DeleteId)"],
                  "/products_for_machines":[f"UPDATE products_for_machines SET Exist = 0 WHERE Id = (DeleteId)", f"SELECT Machine, ProductId, TimeStart FROM products_for_machines WHERE Id = (DeleteId)"],
                  "/products":[f"UPDATE products SET Exist = 0 WHERE Id = (DeleteId)", f"SELECT ShortName, Artikel FROM products WHERE Id = (DeleteId)"],
                  "/worker_shifts":[f"DELETE FROM workers_shifts WHERE Id = (DeleteId)", f"SELECT ShiftStart, ShiftEnd, ShiftTime FROM workers_shifts WHERE Id = (DeleteId)"],
                  "/worker_shift_gloves":[f"", f"SELECT Machine, ProductId, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE Id = (DeleteId)"],
                  "/unloadings":[f"DELETE FROM unloadings WHERE Id = (DeleteId)", f"SELECT Name, Stage, Date FROM unloadings WHERE Id = (DeleteId)"],
                  "/unloading_info":[f"DELETE FROM unloadings_info WHERE Id = (DeleteId)", f"SELECT ProductId, Pair FROM unloadings_info WHERE Id = (DeleteId)"]}


LastRequests = {'/workers': None, '/worker_shifts': None, '/worker_shift_gloves': None, '/worker_shift_machine_gloves': None, '/products_for_machines': None, '/stage_machines_gloves': None, '/products': None, '/product_info': None, '/comings': None, '/coming_info': None, '/schedule': None, '/downloaded_schedule': None, "/salary": None, "/shifts_salary":None, "/salary_set":None, "/unloadings":None, "/unloading_info":None}


def AppScreen(Page):
    Page.title = "Підрахунок запасів рукавичок"
    Page.theme_mode = flet.ThemeMode.LIGHT
    Page.window_width = 1100
           

    def ChangeRoute(RouteEvent):
        global SalaryEditMonth


#Global------------------------------------------------------------------------------------------------------------------------


        def CreateTable(InfoList, Exist, Id, Edit, Delete, ShowRoute, EditRoute, Table, RouteTableColumns, TablePage):
            if Table in ["/worker_shift_machine_gloves", "/stage_machines_gloves", "/product_info", "/coming_info"]:
                AllName=RouteTableColumns[RouteTableColumns.index("Pair")-1]
                RouteRows = [
                    flet.DataRow(
                        cells=[
                            flet.DataCell(
                                flet.Text('Всього:' if RouteTableColumn==AllName else (f"{sum(RouteElementInfo['Pair'] for RouteElementInfo in InfoList)} пар" if RouteTableColumn=="Pair" else ''), size=15, color="#306cae")
                            ) for RouteTableColumn in RouteTableColumns
                        ],
                        color='#fdfcff'
                    )
                ]
            elif Table in ["/worker_shift_gloves"]:
                AllName=RouteTableColumns[RouteTableColumns.index("Pair")-1]
                RouteRows = [
                    flet.DataRow(
                        cells=[
                            flet.DataCell(
                                flet.Text('Всього:' if RouteTableColumn==AllName else (f"{sum(RouteElementInfo['Pair'] for RouteElementInfo in InfoList)} пар" if RouteTableColumn=="Pair" else ''), size=15, color="#306cae")
                            ) for RouteTableColumn in RouteTableColumns+['Edit', 'Delete']
                        ],
                        color='#fdfcff'
                    )
                ]
            elif Table in ["/shifts_salary"]:
                AllName=RouteTableColumns[RouteTableColumns.index("Salary")-3]
                RouteRows = [
                    flet.DataRow(
                        cells=[
                            flet.DataCell(
                                flet.Text('Всього:' if RouteTableColumn==AllName else (f"{sum(RouteElementInfo['Salary'] for RouteElementInfo in InfoList)} грн" if RouteTableColumn=="Salary" else (f"{sum(RouteElementInfo['ShiftPair'] for RouteElementInfo in InfoList)} пар" if RouteTableColumn=="ShiftPair" else (f"{sum(divmod(sum(int(t.split('годин')[0]) * 60 + int(t.split('годин')[1].split('хвилин')[0]) for t in [RouteElementInfo['ShiftTime'] for RouteElementInfo in InfoList]), 60))}годин {divmod(sum(int(t.split('годин')[0]) * 60 + int(t.split('годин')[1].split('хвилин')[0]) for t in [RouteElementInfo['ShiftTime'] for RouteElementInfo in InfoList]), 60)[1]}хвилин" if RouteTableColumn=="ShiftTime" else ''))), size=15, color="#306cae")
                            ) for RouteTableColumn in RouteTableColumns+['Edit']
                        ],
                        color='#fdfcff'
                    )
                ]
            else:
                RouteRows = []

            if Table != '/salary_set':
                PageCounts=math.ceil(len(InfoList)/200)+1
                PageRow.controls=[flet.ElevatedButton(text=str(PageNum), bgcolor="#dde1e7", disabled=True) if TablePage==PageNum else flet.OutlinedButton(text=str(PageNum), on_click=lambda RouteEvent, PageNum=PageNum: ChangePage(PageNum, Table)) for PageNum in range(1, PageCounts)]
                InfoList=InfoList[200*TablePage-200:200*TablePage]
            
            for RouteElementInfo in InfoList:
                RouteCell=[]
                RouteActive=RouteElementInfo['Active'] if 'Active' in RouteElementInfo else False
                RouteId=RouteElementInfo[Id] if Id != None else None
                if Table=='/comings': RouteExist=RouteElementInfo["Pair"] > 0
                else: RouteExist=RouteElementInfo[Exist] if Exist != None else True

                for RouteColumn, RouteElement in RouteElementInfo.items():
                    if RouteColumn not in [Exist, Id, 'Active', "ShiftId"]:
                        ColumnUnits={"Pair":"пар", "Sort": "сорт", "Salary":"грн"}
                        if RouteColumn=="Password":
                            PasswordText=flet.Text("•••••", selectable=True, size=15, color="#306cae")
                            RouteCell.append(
                                flet.DataCell(
                                    flet.Row(controls=[PasswordText, flet.IconButton(icon=flet.icons.KEY_OUTLINED, icon_size=24, on_click=lambda TapEvent, Password=RouteElement, PasswordText=PasswordText: ShowHidePassword(True, PasswordText, Password, TapEvent.control))]),
                                    on_tap=lambda TapEvent, ShowId=RouteId: ShowTableData(ShowId, ShowRoute)
                                )
                            )
                        elif ShowRoute!=None and RouteId!=None:
                            RouteCell.append(
                                flet.DataCell(
                                    flet.Text(f"{RouteElement} {ColumnUnits[RouteColumn]}" if RouteColumn in ColumnUnits else str(RouteElement).replace("година", "г.").replace("годин", "г.").replace("хвилина", "хв.").replace("хвилин", "хв."), selectable=True, size=15, color="#306cae"),
                                    on_tap=lambda TapEvent, ShowId=RouteId: ShowTableData(ShowId, ShowRoute)
                                )
                            )
                        elif Table=='/salary':
                            RouteCell.append(
                                flet.DataCell(
                                    flet.Text(RouteElement, selectable=True, size=15, color="#306cae"),
                                    on_tap=lambda TapEvent, RouteElement=RouteElement: ShowTableData(RouteElement, '/shifts_salary')
                                )
                            )
                        elif Table=='/salary_set' and RouteColumn in ["ShiftPlan", "PairPrice"]:
                            RouteCell.append(
                                flet.DataCell(
                                    flet.TextField(value=RouteElement, suffix_text="пар" if RouteColumn=="ShiftPlan" else "грн", input_filter=flet.InputFilter('^[0-9]*') if RouteColumn=="ShiftPlan" else flet.InputFilter('^[0-9.]*'), height=40)
                                )
                            )
                        else:
                            RouteCell.append(
                                flet.DataCell(
                                    flet.Text(f"{RouteElement} {ColumnUnits[RouteColumn]}" if RouteColumn in ColumnUnits else str(RouteElement).replace("година", "г.").replace("годин", "г.").replace("хвилина", "хв.").replace("хвилин", "хв."), selectable=True, size=15, color="#306cae")
                                )
                            )
                
                if Table == "/salary":
                    RouteCell.append(
                        flet.DataCell(
                            flet.IconButton(
                                icon=flet.icons.DOWNLOAD,
                                icon_size=25,
                                on_click=lambda TapEvent, Date=RouteElementInfo["Month"]: MonthSalaryDownload(Date.split(' ')[0], Date.split(' ')[1])
                            )
                        )
                    )

                    RouteCell.append(
                        flet.DataCell(
                            flet.IconButton(
                                icon=flet.icons.EDIT,
                                icon_size=25,
                                on_click=lambda TapEvent, EditMonth=RouteElementInfo["Month"]: EditSalaryMonthInfo(EditMonth)
                            )
                        )
                    )
                
                if Table == "/shifts_salary":
                    RouteCell.append(
                        flet.DataCell(
                            flet.IconButton(
                                icon=flet.icons.EDIT,
                                icon_size=25,
                                on_click=lambda TapEvent, EditShift=RouteElementInfo["ShiftId"]: EditSalaryShiftInfo(EditShift)
                            )
                        )
                    )

                if Table == "/worker_shifts":
                    RouteCell.append(
                        flet.DataCell(
                            flet.IconButton(
                                icon=flet.icons.DIRECTIONS_CAR_FILLED_ROUNDED,
                                icon_size=25,
                                on_click=lambda TapEvent, RouteElement=RouteId: ShowTableData(RouteElement, '/workers_shifts_machines_adjusts')
                            )
                        )
                    )

                if Edit!=None and RouteExist==True or Table=='/comings':
                    RouteCell.append(
                        flet.DataCell(
                            flet.IconButton(
                                icon=flet.icons.EDIT,
                                icon_size=25,
                                on_click=lambda TapEvent, EditId=RouteId: ToEdit(EditId, EditRoute)
                            )
                        )
                    )
                elif Edit!=None and RouteExist==False:
                    RouteCell.append(
                        flet.DataCell(
                            flet.Text("", size=15, color="#306cae"),
                            on_tap=lambda TapEvent, ShowId=RouteId: ShowTableData(ShowId, ShowRoute)
                        )
                    )

                if Delete!=None and RouteExist==True:
                    RouteCell.append(
                        flet.DataCell(
                            flet.IconButton(
                                icon=flet.icons.CANCEL_ROUNDED,
                                icon_size=25,
                                on_click=lambda TapEvent, DeleteId=RouteId, Table=Table: DeleteInfo(DeleteId, Table)
                            )
                        )
                    )
                elif Delete!=None and RouteExist==False:
                    RouteCell.append(
                        flet.DataCell(
                            flet.Text("", size=15, color="#306cae"),
                            on_tap=lambda TapEvent, ShowId=RouteId: ShowTableData(ShowId, ShowRoute)
                        )
                    )

                if Table == "/schedule":
                    if RouteActive==0:
                        RowColor='#fc8e8e'
                    elif RouteActive==1:
                        RowColor = '#facccc'
                    elif RouteActive==2:
                        RowColor = '#1fc85b'

                    RouteRows.append(
                        flet.DataRow(
                            cells=RouteCell,
                            color=RowColor
                        )
                    )

                else:
                    RouteRows.append(
                        flet.DataRow(
                            cells=RouteCell,
                            color='#1fc85b' if RouteActive==True else ('#fdfcff' if RouteExist == True else '#d1d1d2')
                        )
                    )
                    
            RouteTable.rows = RouteRows
            Page.update()


        def ChangePageStatus(RouteColumns, RouteTableColumns, Table, AddStatus, FilterInput, FilterColumn, WithDeletedCheckbox=None, Exist=None, Id=None, Edit=None, Delete=None, ShowRoute=None, EditRoute=None, SortColumnIndex=None):
            RouteTable.rows=[]
            Page.update()
            

            if AddStatus=="Normal" and CurrentPageStatus[Table]==None and LastRequests[Table]!=None:
                CurrentPageStatus[Table]=LastRequests[Table][3]
                SortColumnIndex=LastRequests[Table][0]
                FilterInput.value=LastRequests[Table][1]
                FilterColumn.value=LastRequests[Table][2]
                if "Sort" in CurrentPageStatus[Table]:
                    AddStatus = "Sort"
            else:
                Statuses = {"Sort":0, "Filter":2, "WithDeleted":3}

                CurrentPageStatus[Table]={Statuses[Status]:Status for Status in str(CurrentPageStatus[Table]).split(" ") if Status in Statuses}

                if AddStatus == "Normal": CurrentPageStatus[Table]={}
                elif Statuses[AddStatus] in CurrentPageStatus[Table] and Statuses[AddStatus]==0 and RouteTable.sort_column_index==SortColumnIndex and RouteTable.sort_ascending==False: CurrentPageStatus[Table][1]="Sort"
                elif Statuses[AddStatus] in CurrentPageStatus[Table] and Statuses[AddStatus]==3: del CurrentPageStatus[Table][Statuses[AddStatus]]
                else: CurrentPageStatus[Table][Statuses[AddStatus]]=AddStatus

                if RouteTable.sort_ascending == True and AddStatus not in ["Sort", "Normal"]: CurrentPageStatus[Table][1]="Sort"

                CurrentPageStatus[Table]=' '.join([CurrentPageStatus[Table][StatusIndex] for StatusIndex in sorted(CurrentPageStatus[Table])])
                
                if CurrentPageStatus[Table]=='': CurrentPageStatus[Table]='Normal'

            if WithDeletedCheckbox!=None: WithDeletedCheckbox.value=True if "WithDeleted" in CurrentPageStatus[Table] else False

            if AddStatus=="Sort" or CurrentPageStatus[Table]=='Sort' and AddStatus!="WithDeleted" or CurrentPageStatus[Table]=='Sort Sort' and AddStatus!="WithDeleted":RouteTable.sort_column_index = SortColumnIndex 
            elif 'Sort' not in CurrentPageStatus[Table]: RouteTable.sort_column_index = None

            RouteTable.sort_ascending = True if 'Sort Sort' in CurrentPageStatus[Table] else False

            if CurrentPageStatus[Table]=="Normal":
                RouteTable.sort_column_index = None
                FilterColumn.value = ''
                FilterInput.value = None


            OpenDB()
            ChangeStatusRequest = PageRequests[Table][CurrentPageStatus[Table].replace('Sort Sort', 'Sort').replace(' ', '')].replace("(FilterColumn)", str(FilterColumn.value).replace('None', '')).replace("(FilterValue)", "%"+FilterInput.value.replace(" ", "%")+"%" if FilterColumn.value != 'Machine' else FilterInput.value).replace("(SortDirection)", 'DESC' if RouteTable.sort_ascending==True else 'ASC').replace("(SortColumn)", RouteTableColumns[int(RouteTable.sort_column_index)] if RouteTable.sort_column_index not in [None, ''] else '')
            
            if Table in ["/worker_shifts", "/worker_shift_gloves", "/worker_shift_machine_gloves", "/product_info", "/coming_info", "/unloading_info"]: ChangeStatusRequest = ChangeStatusRequest.replace("(ActiveId)", str(ActiveId))
            elif Table in ["/stage_machines_gloves"]: ChangeStatusRequest = ChangeStatusRequest.replace("(ActiveStage)", str(ActiveStage))
            elif Table in ["/salary"]:
                DBCursor.execute(f"SELECT MIN(Id) FROM workers_shifts")
                MinShiftId=DBCursor.fetchone()[0]
                ChangeStatusRequest = ChangeStatusRequest.replace("(MinShiftId)", str(MinShiftId))
                
                DBCursor.execute(f"SELECT WorkerId, ShiftPlan, PairPrice FROM salary_info")
                SalaryInfo={WorkerId:(ShiftPlan, PairPrice) for WorkerId, ShiftPlan, PairPrice in DBCursor.fetchall()}

                DBCursor.execute(f"SELECT ShiftId FROM salary")
                WorkersShiftsSalary = [ShiftId[0] for ShiftId in DBCursor.fetchall()]
                
                DBCursor.execute(f"SELECT Id, ShiftStart, WorkerId FROM workers_shifts WHERE ShiftEnd!='?'")
                WorkersShifts = DBCursor.fetchall()
                for ShiftId, ShiftStart, WorkerId in WorkersShifts:
                    if ShiftId not in WorkersShiftsSalary and WorkerId in SalaryInfo:
                        DBCursor.execute(f"INSERT INTO salary VALUES ({ShiftId}, '{ShiftStart}', {WorkerId}, {SalaryInfo[WorkerId][0]}, {SalaryInfo[WorkerId][1]})")
                DBConnector.commit()


            DBCursor.execute(ChangeStatusRequest)
            RouteInfo = DBCursor.fetchall()

            if Table=='/workers':
                ActiveRouteInfo=[]
                for RouteElementInfo in RouteInfo:
                    DBCursor.execute(f"SELECT * FROM workers_shifts WHERE WorkerId = {RouteElementInfo[0]} AND ShiftEnd='?' ") 
                    if len(DBCursor.fetchall())!=0:
                        ActiveRouteInfo.append((RouteElementInfo[0], RouteElementInfo[1], RouteElementInfo[2], RouteElementInfo[3], RouteElementInfo[4], True))
                    else:
                        ActiveRouteInfo.append((RouteElementInfo[0], RouteElementInfo[1], RouteElementInfo[2], RouteElementInfo[3], RouteElementInfo[4], False))
                RouteInfo=ActiveRouteInfo
                RouteColumns.append('Active')
            if Table=='/worker_shifts':
                ActiveRouteInfo=[]

                for RouteElementInfo in RouteInfo:
                    if RouteElementInfo[2]!='?':
                        MachinesAdjustments=eval(RouteElementInfo[-1])
                        ShiftEnd=RouteElementInfo[2]
                        ShiftTime = int(RouteElementInfo[3].split(' ')[0])*3600+int(RouteElementInfo[3].split(' ')[2])*60
                        MachinesAdjustmentsInfo = {Machine: 0 for Machine in MachinesAdjustments[0][0].keys()}
                        for MachinesAdjustmentsIndex, MachinesStates in enumerate(MachinesAdjustments):
                            MachineTimeWorkStart = datetime.datetime.strptime(MachinesStates[1], '%Y-%m-%d %H:%M:%S.%f')
                            MachineTimeWorkEnd = datetime.datetime.strptime(MachinesAdjustments[MachinesAdjustmentsIndex+1][1], '%Y-%m-%d %H:%M:%S.%f') if MachinesAdjustmentsIndex != len(MachinesAdjustments) - 1 else (datetime.datetime.strptime(ShiftEnd, '%d.%m.%Y %H:%M') if ShiftEnd != '?' else datetime.datetime.now())
                            WorkTime = (MachineTimeWorkEnd - MachineTimeWorkStart).total_seconds()
                            for Machine, MachinesState in MachinesStates[0].items():
                                if MachinesState=="0":
                                    MachinesAdjustmentsInfo[Machine] = MachinesAdjustmentsInfo[Machine]+WorkTime

                        Hours, Minutes = divmod(max(MachinesAdjustmentsInfo.values()) // 60, 60)
                        ActiveTime = f"{int(Hours)} годин{'a' if Hours==1 else ''} {int(Minutes)} хвилин{'a' if Minutes==1 else ''}"
                        Hours, Minutes = divmod((ShiftTime-max(MachinesAdjustmentsInfo.values())) // 60, 60)
                        DeactiveTime = f"{int(Hours)} годин{'a' if Hours==1 else ''} {int(Minutes)} хвилин{'a' if Minutes==1 else ''}"
                    else:
                        ActiveTime="?"
                        DeactiveTime="?"
                    ActiveRouteInfo.append((RouteElementInfo[0], RouteElementInfo[1], RouteElementInfo[2], RouteElementInfo[3], ActiveTime, DeactiveTime))
                RouteInfo=ActiveRouteInfo
            elif Table=="/worker_shift_gloves":
                DBCursor.execute("SELECT Id, FullName FROM products")
                Products={Id: FullName for Id, FullName in DBCursor.fetchall()}

                DBCursor.execute(f"""SELECT ShiftStart, ShiftEnd FROM workers_shifts WHERE Id = {SecondaryActiveId}""")
                ShiftStart, ShiftEnd = DBCursor.fetchone()
                RouteInfo=[(Id, Machine, Products[ProductId], Sort, Pair, AddDate) for Id, Machine, ProductId, Sort, Pair, AddDate in RouteInfo if datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M") <= datetime.datetime.strptime(AddDate, "%d.%m.%Y %H:%M") <= (datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M") if ShiftEnd!='?' else datetime.datetime.now())]
            elif Table=="/stage_machines_gloves":
                DBCursor.execute("SELECT Id, ShortName FROM products")
                Products={Id: FullName for Id, FullName in DBCursor.fetchall()}

                RouteInfo=[(Machine, Products[ProductId], Sort, Pair, AddDate) for Machine, ProductId, Sort, Pair, AddDate in RouteInfo]
            elif Table=="/products_for_machines":
                DBCursor.execute("SELECT Id, ShortName FROM products")
                Products={Id: FullName for Id, FullName in DBCursor.fetchall()}

                RouteInfo=[(Id, Machine, Products[ProductId], TimeStart, TimeEnd, Exist) for Id, Machine, ProductId, TimeStart, TimeEnd, Exist in RouteInfo]
            elif Table=="/worker_shift_machine_gloves":
                DBCursor.execute(f"""SELECT ShiftStart, ShiftEnd FROM workers_shifts WHERE Id = {SecondaryActiveId}""")
                ShiftStart, ShiftEnd = DBCursor.fetchone()
                RouteInfo=[(Machine, Sort, Pair) for Machine, Sort, Pair, AddDate in RouteInfo if datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M") <= datetime.datetime.strptime(AddDate, "%d.%m.%Y %H:%M") <= (datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M") if ShiftEnd!='?' else datetime.datetime.now())]
                RouteInfo = [(MachineSort[0], MachineSort[1], sum(Pair for Machine, Sort, Pair in RouteInfo if (Machine, Sort) == MachineSort)) for MachineSort in {(Machine, Sort) for Machine, Sort, Pair in RouteInfo}]
                if not RouteTable.sort_column_index: RouteInfo.sort(key=lambda Column: Column[0], reverse=RouteTable.sort_ascending)
                else: RouteInfo.sort(key=lambda Column: Column[RouteTable.sort_column_index], reverse=RouteTable.sort_ascending)
            elif Table=="/schedule":
                RouteInfo=[(WorkerId, ShiftStart, ShiftEnd) for WorkerId, ShiftStart, ShiftEnd in RouteInfo if datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M")<datetime.datetime.now()]
                DBCursor.execute(f"""SELECT Id, Name, Stage FROM workers""")
                Workers = {Id:(Name, Stage) for Id, Name, Stage in DBCursor.fetchall()}

                DBCursor.execute(f"""SELECT WorkerId, ShiftStart, ShiftEnd, ShiftTime FROM workers_shifts WHERE ShiftEnd!='?'""")
                WorkersShifts=DBCursor.fetchall()
                WorkersShiftsDict = {(WorkerId, ShiftStart, ShiftEnd): ActiveShiftTime for WorkerId, ShiftStart, ShiftEnd, ActiveShiftTime in WorkersShifts}
                WorkersShiftsDate = {(WorkerId, datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M").date(), datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M").date()):(ShiftStart, ShiftEnd, ActiveShiftTime) for WorkerId, ShiftStart, ShiftEnd, ActiveShiftTime in WorkersShifts}

                MinuteOffsets = [datetime.timedelta(minutes=delta) for delta in range(-10, 11)]

                StatusRouteInfo = []
                for WorkerId, ShiftStart, ShiftEnd in RouteInfo:
                    ShiftStartParsed = datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M")
                    ShiftEndParsed = datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M")
                    Minutes = (ShiftEndParsed-ShiftStartParsed).total_seconds() / 60
                    Hours = int(Minutes // 60)
                    Minutes = int(Minutes % 60)
                    ShiftsTimeParsed = f"{Hours} {'годин' if Hours != 1 else 'година'} {Minutes} {'хвилин' if Minutes != 1 else 'хвилина'}"
                    ShiftFound = False
                    for DeltaShiftStart in MinuteOffsets:
                        for DeltaShiftEnd in MinuteOffsets:
                            AdjustedShiftStart = (ShiftStartParsed + DeltaShiftStart).strftime("%d.%m.%Y %H:%M")
                            AdjustedShiftEnd = (ShiftEndParsed + DeltaShiftEnd).strftime("%d.%m.%Y %H:%M")
                            if (WorkerId, AdjustedShiftStart, AdjustedShiftEnd) in WorkersShiftsDict:
                                StatusRouteInfo.append((Workers[WorkerId][0], Workers[WorkerId][1], AdjustedShiftStart, AdjustedShiftEnd, f'{WorkersShiftsDict[(WorkerId, AdjustedShiftStart, AdjustedShiftEnd)]}({ShiftsTimeParsed})', 2))
                                ShiftFound = True
                                break
                        if ShiftFound: break
                    if not ShiftFound and (WorkerId, ShiftStartParsed.date(), ShiftEndParsed.date()) in WorkersShiftsDate:
                        WorkerShiftInfo=WorkersShiftsDate[(WorkerId, ShiftStartParsed.date(), ShiftEndParsed.date())]
                        StatusRouteInfo.append((Workers[WorkerId][0], Workers[WorkerId][1], f'{ShiftStart}({WorkerShiftInfo[0]})', f'{ShiftEnd}({WorkerShiftInfo[1]})', f'{WorkerShiftInfo[2]}({ShiftsTimeParsed})', 1))
                    elif not ShiftFound:
                        StatusRouteInfo.append((Workers[WorkerId][0], Workers[WorkerId][1], ShiftStart, ShiftEnd, f"0 годин 0 хвилин({ShiftsTimeParsed})", 0))
                RouteInfo=StatusRouteInfo
                RouteColumns.append('Active')

                if "Filter" in CurrentPageStatus[Table]:
                    RouteInfo=[RouteInfoElement for RouteInfoElement in RouteInfo if fnmatch.fnmatch(RouteInfoElement[RouteColumns.index(FilterColumn.value)].lower(), f'*{FilterInput.value.lower().replace(" ", "*")}*')]
                if "Sort" in CurrentPageStatus[Table]:
                    RouteInfo.sort(key=lambda Column: Column[RouteTable.sort_column_index], reverse=RouteTable.sort_ascending)
            elif Table=="/downloaded_schedule":
                DBCursor.execute(f"""SELECT Id, Name, Stage FROM workers""")
                Workers = {Id:(Name, Stage) for Id, Name, Stage in DBCursor.fetchall()}
                RouteInfo=[(Workers[WorkerId][0], Workers[WorkerId][1], ShiftStart, ShiftEnd) for WorkerId, ShiftStart, ShiftEnd in RouteInfo]
            elif Table=="/salary":
                MonthNamesByNumber = {1: "Січень", 2: "Лютий", 3: "Березень", 4: "Квітень", 5: "Травень", 6: "Червень", 7: "Липень", 8: "Серпень", 9: "Вересень", 10: "Жовтень", 11: "Листопад", 12: "Грудень"}
                MinMonthAndYear=datetime.datetime.strptime(RouteInfo[0][0], '%d.%m.%Y %H:%M')
                MinMonth, MinYear = MinMonthAndYear.month, MinMonthAndYear.year
                
                MaxMonthAndYear=datetime.datetime.now()
                MaxMonth, MaxYear = MaxMonthAndYear.month, MaxMonthAndYear.year
                RouteInfo = [[
                    (MinMonth + i - 1) % 12 + 1,
                    MinYear + (MinMonth + i - 1) // 12
                ] for i in range((MaxYear - MinYear) * 12 + MaxMonth - MinMonth + 1)]
                RouteInfo.reverse()

                if "Filter" in CurrentPageStatus[Table]:
                    RouteInfo=[RouteInfoElement for RouteInfoElement in RouteInfo if fnmatch.fnmatch(f'{MonthNamesByNumber[RouteInfoElement[0]]} {RouteInfoElement[1]}'.lower(), f'*{FilterInput.value.lower().replace(" ", "*")}*')]

                if "Sort" in CurrentPageStatus[Table]:
                    RouteInfo.sort(key=lambda Column: (Column[1], Column[0]), reverse=RouteTable.sort_ascending)

                RouteInfo = [(f'{MonthNamesByNumber[RouteInfoElement[0]]} {RouteInfoElement[1]}',) for RouteInfoElement in RouteInfo]
            elif Table=='/shifts_salary':
                MonthNumbersByNames = {'Січень': "01", 'Лютий': "02", 'Березень': "03", 'Квітень': "04", 'Травень': "05", 'Червень': "06", 'Липень': "07", 'Серпень': "08", 'Вересень': "09", 'Жовтень': "10", 'Листопад': "11", 'Грудень': "12"}

                DBCursor.execute(f"SELECT Id, ShiftStart, ShiftEnd, ShiftTime FROM workers_shifts WHERE ShiftStart LIKE '%{MonthNumbersByNames[ActiveMonth.split(' ')[0]]}.{ActiveMonth.split(' ')[1]}%'")
                WorkersShiftsInfo = {Id:(ShiftStart, ShiftEnd, ShiftTime) for Id, ShiftStart, ShiftEnd, ShiftTime in DBCursor.fetchall()}

                DBCursor.execute("SELECT Id, Name, Stage FROM workers")
                WorkersNamesInfo = {Id:(Name, Stage) for Id, Name, Stage in DBCursor.fetchall()}

                DBCursor.execute("SELECT WorkerId, Pair, AddDate FROM workers_gloves_quantity WHERE Sort=1")
                WorkersShiftsPairsInfo = DBCursor.fetchall()

                ActiveRouteInfo=[]
                for ShiftId, WorkerId, ShiftPlan, PairPrice in RouteInfo:
                    if ShiftId in WorkersShiftsInfo:
                        ShiftPair=sum(Pair for ShiftWorkerId, Pair, AddDate in WorkersShiftsPairsInfo if ShiftWorkerId==WorkerId and datetime.datetime.strptime(WorkersShiftsInfo[ShiftId][0], "%d.%m.%Y %H:%M") <= datetime.datetime.strptime(AddDate, "%d.%m.%Y %H:%M") <= (datetime.datetime.strptime(WorkersShiftsInfo[ShiftId][1], "%d.%m.%Y %H:%M")))
                        Salary=int((ShiftPair/ShiftPlan)*PairPrice*ShiftPair) if ShiftPlan!=0 else 0
                        ActiveRouteInfo.append((ShiftId, WorkersNamesInfo[WorkerId][0], WorkersNamesInfo[WorkerId][1], WorkersShiftsInfo[ShiftId][0], WorkersShiftsInfo[ShiftId][1], WorkersShiftsInfo[ShiftId][2], ShiftPair, Salary))

                RouteInfo = ActiveRouteInfo
                RouteColumns = ['ShiftId']+RouteColumns

                if "Filter" in CurrentPageStatus[Table]:
                    RouteInfo=[RouteInfoElement for RouteInfoElement in RouteInfo if fnmatch.fnmatch(str(RouteInfoElement[RouteColumns.index(FilterColumn.value)]).lower(), f'*{FilterInput.value.lower().replace(" ", "*")}*')]
                if "Sort" in CurrentPageStatus[Table]:
                    RouteInfo.sort(key=lambda Column: Column[RouteTable.sort_column_index+1], reverse=RouteTable.sort_ascending)
            elif Table=="/salary_set":
                MonthNumbersByNames = {'Січень': "01", 'Лютий': "02", 'Березень': "03", 'Квітень': "04", 'Травень': "05", 'Червень': "06", 'Липень': "07", 'Серпень': "08", 'Вересень': "09", 'Жовтень': "10", 'Листопад': "11", 'Грудень': "12"}

                if SalaryEditMonth==None:
                    DBCursor.execute("SELECT WorkerId, ShiftPlan, PairPrice FROM salary_info")
                    SalaryInfo = {WorkerId:(ShiftPlan, PairPrice) for WorkerId, ShiftPlan, PairPrice in DBCursor.fetchall()}
                else:
                    DBCursor.execute(f"SELECT WorkerId, ShiftPlan, PairPrice FROM salary WHERE ShiftStart LIKE '%{MonthNumbersByNames[SalaryEditMonth.split(' ')[0]]}.{SalaryEditMonth.split(' ')[1]}%'")
                    SalaryInfo=set(DBCursor.fetchall())
                    SalaryInfo = {WorkerId:(ShiftPlan, PairPrice) for WorkerId, ShiftPlan, PairPrice in SalaryInfo}
                    
                RouteInfo = [(f'{WorkerId}.{WorkerName} ({WorkerStage})', str(SalaryInfo[WorkerId][0]) if WorkerId in SalaryInfo else '', str(SalaryInfo[WorkerId][1]) if WorkerId in SalaryInfo else '') for WorkerId, WorkerName, WorkerStage in RouteInfo]
                RouteInfo.reverse()
            elif Table=="/unloading_info":
                DBCursor.execute("SELECT Id, ShortName FROM products")
                Products={Id: ShortName for Id, ShortName in DBCursor.fetchall()}

                RouteInfo=[(Id, Products[ProductId], Pair) for Id, ProductId, Pair in RouteInfo]

            if SortColumnIndex!=None and "Sort" in CurrentPageStatus[Table] and RouteTableColumns[SortColumnIndex] in ["ShiftStart", "ShiftEnd", "Pauses", "TimeStart", "TimeEnd", "AddDate"]:
                if Exist==None: RouteInfo.sort(key=lambda Column: datetime.datetime.strptime(Column[RouteColumns.index(RouteTableColumns[SortColumnIndex])].split('(')[0].split(',')[0].replace('?', datetime.datetime.now().strftime('%d.%m.%Y %H:%M')) if Column[RouteColumns.index(RouteTableColumns[SortColumnIndex])]!="" else "01.01.2020 00:00", "%d.%m.%Y %H:%M"), reverse=RouteTable.sort_ascending)
                else: RouteInfo.sort(key=lambda Column: (Column[RouteColumns.index(Exist)], datetime.datetime.strptime(Column[RouteColumns.index(RouteTableColumns[SortColumnIndex])].split('(')[0].split(',')[0].replace('?', datetime.datetime.now().strftime('%d.%m.%Y %H:%M')) if Column[RouteColumns.index(RouteTableColumns[SortColumnIndex])]!="" else "01.01.2020 00:00", "%d.%m.%Y %H:%M")), reverse=RouteTable.sort_ascending)
            if SortColumnIndex!=None and "Sort" in CurrentPageStatus[Table] and RouteTableColumns[SortColumnIndex] in ["ShiftTime", "ActiveShiftTime", "PausesTime"]:
                if Exist==None: RouteInfo.sort(key=lambda Column: float(".".join(Column[RouteColumns.index(RouteTableColumns[SortColumnIndex])].split()[0::2][:2]).replace("?", "0")), reverse=RouteTable.sort_ascending)
                else: RouteInfo.sort(key=lambda Column: (Column[RouteColumns.index(Exist)], float(".".join(Column[RouteColumns.index(RouteTableColumns[SortColumnIndex])].split()[0::2][:2]).replace("?", "0"))), reverse=RouteTable.sort_ascending)
            
            CloseDB()

            InfoList=[{RouteColumns[RouteElementId]:RouteElement for RouteElementId, RouteElement in enumerate(RouteElements)} for RouteElements in RouteInfo]
            LastRequests[Table]=(SortColumnIndex, FilterInput.value, FilterColumn.value, CurrentPageStatus[Table], InfoList, Exist, Id, Edit, Delete, ShowRoute, EditRoute, Table, RouteTableColumns)


            CreateTable(InfoList, Exist, Id, Edit, Delete, ShowRoute, EditRoute, Table, RouteTableColumns, 1)


        def ChangePage(PageNum, Table):
            CreateTable(LastRequests[Table][4], LastRequests[Table][5], LastRequests[Table][6], LastRequests[Table][7], LastRequests[Table][8], LastRequests[Table][9], LastRequests[Table][10], LastRequests[Table][11], LastRequests[Table][12], PageNum)


        def ShowTableData(ShowId, ShowRoute):
            global ActiveId, SecondaryActiveId, ActiveMonth
            if ShowRoute in ["/worker_shift_info", "/workers_shifts_machines_adjusts"]:
                SecondaryActiveId=ShowId
            elif ShowRoute in ["/shifts_salary"]:
                ActiveMonth=ShowId
            else:
                ActiveId=ShowId
            Page.go(ShowRoute)


        def DeleteInfo(Id, DBTable):
            def DeleteAlert(ClickEvent):
                OpenDB()
                if DBTable == "/worker_shift_gloves":
                    DBCursor.execute(f"SELECT ProductId, AddDate, Pair FROM workers_gloves_quantity WHERE Id = {Id}")
                    ProductId, AddDate, Pair = DBCursor.fetchone()

                    DBCursor.execute(f"SELECT Id From products_gloves_quantity WHERE ProductId = {ProductId} AND Pair={Pair} AND AddDate='{AddDate}'")
                    ProductId=max(DBCursor.fetchone())

                    DBCursor.reset()

                    DBCursor.execute(f"DELETE FROM workers_gloves_quantity WHERE Id = {Id}")

                    DBCursor.execute(f"DELETE FROM products_gloves_quantity WHERE Id = {ProductId}")
                elif DBTable == "/unloadings":
                    DBCursor.execute("DELETE FROM unloadings_info WHERE UnloadingId = (DeleteId)".replace("(DeleteId)", str(Id)))
                    DBCursor.execute(DeleteRequests[DBTable][0].replace("(DeleteId)", str(Id)))
                else:
                    DBCursor.execute(DeleteRequests[DBTable][0].replace("(DeleteId)", str(Id)))

                DBConnector.commit()
                CloseDB()

                ExistAlert.open = False
                Page.update()

                CurrentPageStatus[DBTable]=None
                if DBTable == "/workers":
                    ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker")
                elif DBTable == "/products_for_machines":
                    ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "Product", "TimeStart", "TimeEnd"], "/products_for_machines", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine")
                elif DBTable == "/products":
                    ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product")
                elif DBTable == "/worker_shifts":
                    ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info")
                elif DBTable == "/worker_shift_gloves":
                    ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves")
                elif DBTable == "/unloadings":
                    ChangePageStatus(["Id", "Name", "Stage", "Date"], ["Name", "Stage", "Date"], "/unloadings", "Normal", FilterInput, FilterColumn, Id="Id", Delete=True, ShowRoute="/unloading_info")
                elif DBTable == "/unloading_info":
                    ChangePageStatus(["Id", "Product", "Pair"], ["Product", "Pair"], "/unloading_info", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_unloading")

            def CloseAlert(ClickEvent):
                ExistAlert.open = False
                Page.update()

            OpenDB()
            DBCursor.execute(DeleteRequests[DBTable][1].replace("(DeleteId)", str(Id)))
            if DBTable == "/workers":
                DeleteInfo=DBCursor.fetchone()
                DeleteInfo=f"{DeleteInfo[0]} ({DeleteInfo[1]})"
            elif DBTable == "/products_for_machines":
                DeleteInfo=DBCursor.fetchone()
                DBCursor.execute(f"SELECT ShortName FROM products WHERE Id={DeleteInfo[1]}")
                Product=DBCursor.fetchone()[0]
                DeleteInfo=f"з {DeleteInfo[0]} машини ({Product}, {DeleteInfo[2]})"
            elif DBTable == "/products":
                DeleteInfo=DBCursor.fetchone()
                DeleteInfo=f"{DeleteInfo[0]} ({DeleteInfo[1]})"
            elif DBTable == "/worker_shifts":
                DeleteInfo=DBCursor.fetchone()
                DeleteInfo=f"зміну {DeleteInfo[0]} - {DeleteInfo[1]} ({DeleteInfo[2]})"
            elif DBTable == "/worker_shift_gloves":
                DeleteInfo=DBCursor.fetchone()
                DBCursor.execute(f"SELECT ShortName FROM products WHERE Id={DeleteInfo[1]}")
                Product=DBCursor.fetchone()[0]
                DeleteInfo=f"з {DeleteInfo[0]} машини ({Product}, {DeleteInfo[2]}, {DeleteInfo[3]}, {DeleteInfo[4]})"
            elif DBTable == "/unloadings":
                DeleteInfo=DBCursor.fetchone()
                DeleteInfo=f"вивантаження {DeleteInfo[0]} ({DeleteInfo[1]}), {DeleteInfo[2]}"
            elif DBTable == "/unloading_info":
                DeleteInfo=DBCursor.fetchone()
                DBCursor.execute(f"SELECT ShortName FROM products WHERE Id={DeleteInfo[0]}") 
                ShortNameProduct=DBCursor.fetchone()[0]
                DeleteInfo=f"{DeleteInfo[1]} пар {ShortNameProduct}"

            
            
            CloseDB()

            ExistAlert = flet.AlertDialog(
                modal=True,
                title=flet.Text(f"Ви точно хочете видалити {DeleteInfo}!"),
                actions=[
                    flet.ElevatedButton("Відмінити", on_click=CloseAlert),
                    flet.ElevatedButton("Видалити", on_click=DeleteAlert, bgcolor="red", color="white")
                ]
            )
            Page.dialog = ExistAlert
            ExistAlert.open = True
            Page.update()


        def ToEdit(Id, EditRoute):
            global EditId
            EditId=Id
            Page.go(EditRoute)


#------------------------------------------------------------------------------------------------------------------------


#Worker------------------------------------------------------------------------------------------------------------------------


        def ShowHidePassword(Show, PasswordText, Password, ButtonPassword):
            if Show==True:
                PasswordText.value=Password
                ButtonPassword.icon = flet.icons.KEY_OFF_OUTLINED
                ButtonPassword.on_click=lambda TapEvent, Password=Password, PasswordText=PasswordText: ShowHidePassword(False, PasswordText, Password, TapEvent.control)
            else:
                PasswordText.value="•••••"
                ButtonPassword.icon = flet.icons.KEY_OUTLINED
                ButtonPassword.on_click=lambda TapEvent, Password=Password, PasswordText=PasswordText: ShowHidePassword(True, PasswordText, Password, TapEvent.control)
            Page.update()


        def GetReportDate():
            def ContinueReport(ClickEvent):
                try:
                    ReportDateFrom=datetime.datetime.strptime(ReportInputDateFrom.value, "%d.%m.%Y")
                except:
                    def CloseAlert(ClickEvent):
                        NameAlert.open = False
                        Page.dialog = GetReportDateDialog
                        Page.update()

                    NameAlert = flet.AlertDialog(title=flet.Text("Дата з введена невірно!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                    Page.dialog = NameAlert
                    NameAlert.open = True
                    Page.update()
                    return True

                try:
                    ReportDateTo=datetime.datetime.strptime(ReportInputDateTo.value, "%d.%m.%Y")
                except:
                    def CloseAlert(ClickEvent):
                        NameAlert.open = False
                        Page.dialog = GetReportDateDialog
                        Page.update()

                    NameAlert = flet.AlertDialog(title=flet.Text("Дата до введена невірно!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                    Page.dialog = NameAlert
                    NameAlert.open = True
                    Page.update()
                    return True


                if ReportDateFrom>ReportDateTo:
                    def CloseAlert(ClickEvent):
                        NameAlert.open = False
                        Page.dialog = GetReportDateDialog
                        Page.update()

                    NameAlert = flet.AlertDialog(title=flet.Text("Дата з більша за дату до!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                    Page.dialog = NameAlert
                    NameAlert.open = True
                    Page.update()
                else:
                    ReportDate=(ReportDateFrom.date(), ReportDateTo.date())
                    GetReportDateDialog.open = False
                    Page.update()

                    GetReportStage(ReportDate)


            def CloseGetReportDateDialog(ClickEvent):
                GetReportDateDialog.open = False
                Page.update()

            ReportInputDateFrom = flet.TextField(label=f"З ({datetime.datetime.now().strftime('%d.%m.%Y')})", value=(datetime.datetime.now()- datetime.timedelta(weeks=1)).strftime('%d.%m.%Y'))
            ReportInputDateTo = flet.TextField(label=f"До ({datetime.datetime.now().strftime('%d.%m.%Y')})", value=datetime.datetime.now().strftime('%d.%m.%Y'))
            GetReportDateDialog = flet.AlertDialog(
                modal=True,
                title=flet.Text("Вибери проміжок дати звіту"), 
                content=flet.Column(controls=[ReportInputDateFrom, ReportInputDateTo], height=110),
                actions=[flet.TextButton("Відмінити", on_click=CloseGetReportDateDialog), flet.TextButton("Продовжити", on_click=ContinueReport)],
                actions_alignment=flet.MainAxisAlignment.END,
            )
            
            Page.dialog = GetReportDateDialog
            GetReportDateDialog.open = True
            Page.update()


        def GetReportStage(ReportDate):
            def ContinueReport(ClickEvent):
                ReportStage = ReportInputStage.value
                GetReportStageDialog.open = False
                Page.update()

                FolderPicker.on_result = lambda GetReportEvent: GetReport(GetReportEvent, ReportDate, ReportStage)
                FolderPicker.get_directory_path(dialog_title="Вибери папку для збереження звіту")
                

            def CloseGetReportStageDialog(ClickEvent):
                GetReportStageDialog.open = False
                Page.update()

            ReportInputStage = flet.Dropdown(label="Етап", hint_text="Вибери етап", options=[flet.dropdown.Option("В'язання"), flet.dropdown.Option('ПВХ'), flet.dropdown.Option('Оверлок'), flet.dropdown.Option('Упаковка')], value="В'язання", width=400)
            GetReportStageDialog = flet.AlertDialog(
                modal=True,
                title=flet.Text("Вибери етап звіту"), 
                content=flet.Column(controls=[ReportInputStage], height=60),
                actions=[flet.TextButton("Відмінити", on_click=CloseGetReportStageDialog), flet.TextButton("Продовжити", on_click=ContinueReport)],
                actions_alignment=flet.MainAxisAlignment.END,
            )
            
            Page.dialog = GetReportStageDialog
            GetReportStageDialog.open = True
            Page.update()


        def GetReport(GetReportEvent, ReportDate, ReportStage):
            if GetReportEvent.path != None:
                OpenDB()
                DBCursor.execute(f"""SELECT Id, Name FROM workers WHERE Stage='{ReportStage.replace("'", "''")}'""")
                WorkerInfo=DBCursor.fetchall()
                Workers = f"({', '.join([str(Worker[0]) for Worker in WorkerInfo])})"
                WorkerIdName = {Id:Name for Id, Name in WorkerInfo}

                DBCursor.execute(f"""SELECT WorkerId, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments FROM workers_shifts WHERE WorkerId IN {Workers}""")
                WorkersShifts = DBCursor.fetchall()
                WorkersShiftsInfo=[]
                for WorkerId, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments in WorkersShifts:
                    if datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M").date()>=ReportDate[0] and datetime.datetime.strptime(ShiftEnd.replace("?", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")), "%d.%m.%Y %H:%M").date()<=ReportDate[1]:
                        if ShiftEnd!="?":
                            MachinesAdjustments=eval(MachinesAdjustments)
                            ShiftTimeSeconds = int(ShiftTime.split(' ')[0])*3600+int(ShiftTime.split(' ')[2])*60
                            MachinesAdjustmentsInfo = {Machine: 0 for Machine in MachinesAdjustments[0][0].keys()}
                            for MachinesAdjustmentsIndex, MachinesStates in enumerate(MachinesAdjustments):
                                MachineTimeWorkStart = datetime.datetime.strptime(MachinesStates[1], '%Y-%m-%d %H:%M:%S.%f')
                                MachineTimeWorkEnd = datetime.datetime.strptime(MachinesAdjustments[MachinesAdjustmentsIndex+1][1], '%Y-%m-%d %H:%M:%S.%f') if MachinesAdjustmentsIndex != len(MachinesAdjustments) - 1 else (datetime.datetime.strptime(ShiftEnd, '%d.%m.%Y %H:%M') if ShiftEnd != '?' else datetime.datetime.now())
                                WorkTime = (MachineTimeWorkEnd - MachineTimeWorkStart).total_seconds()
                                for Machine, MachinesState in MachinesStates[0].items():
                                    if MachinesState=="0":
                                        MachinesAdjustmentsInfo[Machine] = MachinesAdjustmentsInfo[Machine]+WorkTime

                            Hours, Minutes = divmod(max(MachinesAdjustmentsInfo.values()) // 60, 60)
                            ActiveTime = f"{int(Hours)} годин{'a' if Hours==1 else ''} {int(Minutes)} хвилин{'a' if Minutes==1 else ''}"
                            Hours, Minutes = divmod((ShiftTimeSeconds-max(MachinesAdjustmentsInfo.values())) // 60, 60)
                            DeactiveTime = f"{int(Hours)} годин{'a' if Hours==1 else ''} {int(Minutes)} хвилин{'a' if Minutes==1 else ''}"
                        else:
                            ActiveTime="?"
                            DeactiveTime="?"
                        WorkersShiftsInfo.append((WorkerId, ShiftStart, ShiftEnd, ShiftTime, ActiveTime, DeactiveTime))

                WorkersShifts = WorkersShiftsInfo

                DBCursor.execute(f"""SELECT Quantity FROM machines_info WHERE Stage='{ReportStage.replace("'", "''")}'""")
                Machines=DBCursor.fetchone()[0]
                
                DBCursor.execute(f"""SELECT WorkerId, Machine, Pair, AddDate FROM workers_gloves_quantity WHERE Sort = 1""")
                WorkersShiftsMachinesGloves = DBCursor.fetchall()
                CloseDB()

                ReportXlsxFile = openpyxl.Workbook()
                ReportXlsxSheet = ReportXlsxFile.active
                ReportXlsxSheet.title = f"Звіт {ReportStage}"

                ReportXlsxSheet.column_dimensions["A"].width = 20

                ReportXlsxSheet["A1"]="Працівник"
                ReportXlsxSheet[f"A1"].font = Font(color = "1E3559", size=14, bold=True)

                ReportXlsxSheet["A2"]="Початок зміни"
                ReportXlsxSheet[f"A2"].font = Font(color = "1E3559", size=14, bold=True)

                ReportXlsxSheet["A3"]="Кінець зміни"
                ReportXlsxSheet[f"A3"].font = Font(color = "1E3559", size=14, bold=True)

                ReportXlsxSheet["A4"]="Час зміни"
                ReportXlsxSheet[f"A4"].font = Font(color = "1E3559", size=14, bold=True)

                ReportXlsxSheet["A5"]="Активний час"
                ReportXlsxSheet[f"A5"].font = Font(color = "1E3559", size=14, bold=True)

                ReportXlsxSheet["A6"]="Деактивний час"
                ReportXlsxSheet[f"A6"].font = Font(color = "1E3559", size=14, bold=True)

                ReportXlsxSheet["A7"]="Машини/Всього"
                ReportXlsxSheet[f"A7"].fill=PatternFill("solid", start_color="006100")
                ReportXlsxSheet[f"A7"].font = Font(color = "C6F0CE", size=14, bold=True)

                MachineCountColumnLetter = get_column_letter(len(WorkersShifts)+2)

                ReportXlsxSheet.column_dimensions[MachineCountColumnLetter].width = 20
                ReportXlsxSheet[f"{MachineCountColumnLetter}7"]="Всього на машину"
                ReportXlsxSheet[f"{MachineCountColumnLetter}7"].fill=PatternFill("solid", start_color="006100")
                ReportXlsxSheet[f"{MachineCountColumnLetter}7"].font = Font(color = "C6F0CE", size=14, bold=True)

                
                for Machine in range(1, Machines+1):
                    ReportXlsxSheet[f"A{Machine+7}"]=str(Machine)
                    ReportXlsxSheet[f"A{Machine+7}"].fill=PatternFill("solid", start_color="C6F0CE")
                    ReportXlsxSheet[f"A{Machine+7}"].font = Font(color = "006100")
                    ReportXlsxSheet[f"A{Machine+7}"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                    MachinePairCount = sum(sum(Pair for WorkerShiftId, MachineShift, Pair, AddDate in WorkersShiftsMachinesGloves if WorkerShiftId==WorkerId and MachineShift==Machine and datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M") <= datetime.datetime.strptime(AddDate, "%d.%m.%Y %H:%M") <= (datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M") if ShiftEnd!='?' else datetime.datetime.now())) for WorkerId, ShiftStart, ShiftEnd, ShiftTime, ActiveTime, DeactiveTime in WorkersShifts)
                    ReportXlsxSheet[f"{MachineCountColumnLetter}{Machine+7}"]=f'{Machine} - {str(MachinePairCount).replace(".0", "")} пар'
                    ReportXlsxSheet[f"{MachineCountColumnLetter}{Machine+7}"].fill=PatternFill("solid", start_color="C6F0CE")
                    ReportXlsxSheet[f"{MachineCountColumnLetter}{Machine+7}"].font = Font(color = "006100")
                    ReportXlsxSheet[f"{MachineCountColumnLetter}{Machine+7}"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                ReportXlsxSheet["A7"].border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
                ReportXlsxSheet[f"A{Machines+7}"].border=Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
                
                for WorkersShiftsIndex, WorkerShift in enumerate(WorkersShifts):
                    WorkerId, ShiftStart, ShiftEnd, ShiftTime, ActiveTime, DeactiveTime=WorkerShift
                    WorkerShiftColumnLetter = get_column_letter(WorkersShiftsIndex+2)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}1"]=WorkerIdName[WorkerId]
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}1"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}1"].font = Font(color = "7B002C", bold=True)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}1"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}2"]=ShiftStart
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}2"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}2"].font = Font(color = "7B002C", bold=True)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}2"].border=Border(left=Side(style='thin'), right=Side(style='thin'))
                    
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}3"]=ShiftEnd
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}3"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}3"].font = Font(color = "7B002C", bold=True)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}3"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}4"]=ShiftTime.replace("година", "г.").replace("годин", "г.").replace("хвилина", "хв.").replace("хвилин", "хв.")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}4"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}4"].font = Font(color = "7B002C", bold=True)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}4"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}5"]=ActiveTime.replace("година", "г.").replace("годин", "г.").replace("хвилина", "хв.").replace("хвилин", "хв.")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}5"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}5"].font = Font(color = "7B002C", bold=True)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}5"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}6"]=DeactiveTime.replace("година", "г.").replace("годин", "г.").replace("хвилина", "хв.").replace("хвилин", "хв.")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}6"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}6"].font = Font(color = "7B002C", bold=True)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}6"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}7"]=str(sum(sum(Pair for WorkerShiftId, MachineShift, Pair, AddDate in WorkersShiftsMachinesGloves if WorkerShiftId==WorkerId and MachineShift==Machine and datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M") <= datetime.datetime.strptime(AddDate, "%d.%m.%Y %H:%M") <= (datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M") if ShiftEnd!='?' else datetime.datetime.now())) for Machine in range(1, Machines+1)))
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}7"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}7"].font = Font(color = "7B002C", bold=True)
                    ReportXlsxSheet[f"{WorkerShiftColumnLetter}7"].border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    ReportXlsxSheet.column_dimensions[WorkerShiftColumnLetter].width = 33

                    for Machine in range(1, Machines+1):
                        MachinePairCount = str(sum(Pair for WorkerShiftId, MachineShift, Pair, AddDate in WorkersShiftsMachinesGloves if WorkerShiftId==WorkerId and MachineShift==Machine and datetime.datetime.strptime(ShiftStart, "%d.%m.%Y %H:%M") <= datetime.datetime.strptime(AddDate, "%d.%m.%Y %H:%M") <= (datetime.datetime.strptime(ShiftEnd, "%d.%m.%Y %H:%M") if ShiftEnd!='?' else datetime.datetime.now())))
                        ReportXlsxSheet[f"{WorkerShiftColumnLetter}{Machine+7}"]=MachinePairCount.replace('.0', '')
                        ReportXlsxSheet[f"{WorkerShiftColumnLetter}{Machine+7}"].fill=PatternFill("solid", start_color="F8FCD0" if WorkersShiftsIndex%2==0 else "F8FCAE")
                        ReportXlsxSheet[f"{WorkerShiftColumnLetter}{Machine+7}"].font = Font(color = "9C5700")
                        ReportXlsxSheet[f"{WorkerShiftColumnLetter}{Machine+7}"].border=Border(left=Side(style='thin'), right=Side(style='thin'))

                        ReportXlsxSheet[f"{WorkerShiftColumnLetter}8"].border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
                        ReportXlsxSheet[f"{WorkerShiftColumnLetter}{Machines+7}"].border=Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

                ReportXlsxFile.save(f"""{GetReportEvent.path}/Звіт по {ReportStage} ({ReportDate[0].strftime("%d.%m.%Y")}-{ReportDate[1].strftime("%d.%m.%Y")}).xlsx""")

                def CloseAlert(ClickEvent):
                    Alert.open = False
                    Page.update()

                Alert = flet.AlertDialog(title=flet.Text(f"""Звіт був збережений тут - {GetReportEvent.path}/Звіт по {ReportStage} ({ReportDate[0].strftime("%d.%m.%Y")}-{ReportDate[1].strftime("%d.%m.%Y")}).xlsx""", selectable=True), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = Alert
                Alert.open = True
                Page.update()


        def EditWorkerShift(RouteEvent):
            ShiftStart=''
            ShiftEnd=''
            try:
                ShiftStart=datetime.datetime.strptime(ShiftStartTextField.value, "%d.%m.%Y %H:%M")
            except:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Початок зміни введений невірно!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()

            try:
                ShiftEnd=datetime.datetime.strptime(ShiftEndTextField.value, "%d.%m.%Y %H:%M")
            except:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Кінець зміни введений невірно!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            if ShiftStart!="" and ShiftEnd!="":
                Minutes = (ShiftEnd - ShiftStart).total_seconds() / 60
                Hours = int(Minutes // 60)
                Minutes = int(Minutes % 60)
                ShiftTime = f"{Hours} {'годин' if Hours != 1 else 'година'} {Minutes} {'хвилин' if Minutes != 1 else 'хвилина'}"
                OpenDB()
                DBCursor.execute(f"""UPDATE workers_shifts SET ShiftStart='{ShiftStart.strftime("%d.%m.%Y %H:%M")}', ShiftEnd='{ShiftEnd.strftime("%d.%m.%Y %H:%M")}', ShiftTime='{ShiftTime}' WHERE Id={EditId}""")
                DBConnector.commit()
                CloseDB()

                Page.go('/worker_shifts')


        def EditWorkerShiftGloves(RouteEvent):
            if str(GlovesCountTextField.value).replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    Alert.open = False
                    Page.update()

                Alert = flet.AlertDialog(title=flet.Text("Кількість не вибрана!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = Alert
                Alert.open = True
                Page.update()
            elif ProductTextField.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Продукт не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"SELECT Id FROM products WHERE ShortName = '{ProductTextField.value}'")
                EditProductId = DBCursor.fetchone()[0]

                DBCursor.execute(f"SELECT ProductId, AddDate, Pair FROM workers_gloves_quantity WHERE Id = {EditId}")
                ProductId, AddDate, Pair = DBCursor.fetchone()

                DBCursor.execute(f"SELECT Id From products_gloves_quantity WHERE ProductId = {ProductId} AND Pair={Pair} AND AddDate='{AddDate}'")
                ProductId=max(DBCursor.fetchone())

                Pair=int(GlovesCountTextField.value)/2

                DBCursor.execute(f"UPDATE workers_gloves_quantity SET Pair = {Pair}, ProductId={EditProductId} WHERE Id = {EditId}")

                DBCursor.execute(f"UPDATE products_gloves_quantity SET Pair = {Pair}, ProductId={EditProductId} WHERE Id = {ProductId}")

                DBConnector.commit()

                CloseDB()

                Page.go('/worker_shift_gloves')


        def CreateWorkersShiftsMachinesAdjusts():
            OpenDB()
            DBCursor.execute(f"""SELECT ShiftEnd, MachinesAdjustments FROM workers_shifts WHERE Id = {SecondaryActiveId}""")
            ShiftEnd, MachinesAdjustments=DBCursor.fetchone()
            MachinesAdjustments=eval(MachinesAdjustments)
            CloseDB()

            MachinesAdjustmentsInfo={Machine:{} for Machine in MachinesAdjustments[0][0].keys()}

            for MachinesAdjustmentsIndex, MachinesStates in enumerate(MachinesAdjustments):
                MachineTimeWorkStart = datetime.datetime.strptime(MachinesStates[1], '%Y-%m-%d %H:%M:%S.%f')
                MachineTimeWorkEnd = datetime.datetime.strptime(
                    MachinesAdjustments[MachinesAdjustmentsIndex+1][1], '%Y-%m-%d %H:%M:%S.%f'
                ) if MachinesAdjustmentsIndex != len(MachinesAdjustments) - 1 else (
                    datetime.datetime.strptime(ShiftEnd, '%d.%m.%Y %H:%M')
                    if ShiftEnd != '?' else datetime.datetime.now()
                )
                WorkTime, WorkText = (MachineTimeWorkEnd - MachineTimeWorkStart).total_seconds(), f"{MachineTimeWorkStart.strftime('%d.%m.%Y %H:%M')} - {MachineTimeWorkEnd.strftime('%d.%m.%Y %H:%M')}"

                for Machines, MachinesState in MachinesStates[0].items():
                    if MachinesState not in MachinesAdjustmentsInfo[Machines]:
                        MachinesAdjustmentsInfo[Machines][MachinesState] = [WorkTime, WorkText]
                    else:
                        MachinesAdjustmentsInfo[Machines][MachinesState][0] += WorkTime
                        MachinesAdjustmentsInfo[Machines][MachinesState][1] += f', {WorkText}'

            for Machine in MachinesAdjustmentsInfo:
                States = []
                for MachineState, MachineStateValue in MachinesAdjustmentsInfo[Machine].items():
                    Hours, Minutes = divmod(MachineStateValue[0] // 60, 60)
                    MachineStateValue[0] = f"{int(Hours)} годин{'a' if Hours==1 else ''} {int(Minutes)} хвилин{'a' if Minutes==1 else ''}"
                    
                    MachineShiftTimes=[]
                    for MachineShiftTime in MachineStateValue[1].split(', '):
                        if len(MachineShiftTimes)>0 and MachineShiftTimes[-1].split(" - ")[1]==MachineShiftTime.split(' - ')[0]: 
                            MachineShiftTimes[-1] = f"""{MachineShiftTimes[-1].split(" - ")[0]} - {MachineShiftTime.split(' - ')[1]}"""
                        else: 
                            MachineShiftTimes.append(MachineShiftTime)
                            
                    MachineShiftTimes=', '.join(MachineShiftTimes)


                    States.append((MachineState, MachineStateValue[0], MachineShiftTimes))
                MachinesAdjustmentsInfo[Machine]=States

            RouteRows = []
            for Machine, StateInfo in MachinesAdjustmentsInfo.items():
                RouteCell=[
                    flet.DataCell(
                        flet.Row(controls=[
                            flet.Text(Machine, size=15, color="#306cae", selectable=True), 
                            flet.Icon(
                                name=flet.icons.DIRECTIONS_CAR_FILLED_ROUNDED,
                                size=25
                            )
                        ])
                    )
                ]
                for State, StateTime, StateTimes in StateInfo:
                    RouteCell.append(flet.DataCell(
                        flet.Column(controls=[
                            flet.Icon(
                                name=[flet.icons.CHECK_CIRCLE_OUTLINE_ROUNDED, flet.icons.HIGHLIGHT_REMOVE_ROUNDED, flet.icons.FLASHLIGHT_OFF_ROUNDED, flet.icons.SETTINGS_OUTLINED][int(State)],
                                size=25,
                                color=["#669c35", "#e32400", "#8d8600", "#00364a"][int(State)]
                            ),
                            flet.Text(["Працює", "Не працює", "Вимкнений", "Немає продукту"][int(State)], size=15, color="#306cae", selectable=True),
                            flet.Text(StateTime, size=15, color="#306cae", selectable=True),
                            flet.Text(StateTimes, size=15, color="#306cae", selectable=True), 
                        ], spacing=5, horizontal_alignment=flet.CrossAxisAlignment.START, width=200)
                    ))
                
                for CellIndex in range(5-len(RouteCell)):
                    RouteCell.append(flet.DataCell(flet.Text('')))

                RouteRows.append(flet.DataRow(cells=RouteCell))

            RouteTable.rows=RouteRows

            Page.update()


        def AddWorker():
            OpenDB()
            DBCursor.execute("SELECT Name, Stage FROM workers")
            NameAndStageList=DBCursor.fetchall()
            CloseDB()
            if NameTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Ім'я не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif StageDropdown.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Етап не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif PasswordTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Пароль не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif (NameTextField.value, StageDropdown.value) in NameAndStageList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Такий працівник вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"SELECT MAX(Id) FROM workers")
                Id=DBCursor.fetchone()[0]
                Id=Id+1 if Id != None else 0
                                 
                DBCursor.execute(f"""INSERT INTO workers VALUES ({Id}, '{NameTextField.value}', '{StageDropdown.value.replace("'", "''")}', '{PasswordTextField.value}', True)""")
                DBConnector.commit()
                CloseDB()

                Page.go('/workers')


        def EditWorker():
            OpenDB()
            DBCursor.execute(f"SELECT Name, Stage FROM workers WHERE Id != {EditId}")
            NameAndStageList=DBCursor.fetchall()
            CloseDB()
            if NameTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Ім'я не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif StageDropdown.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Етап не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif PasswordTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Пароль не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif (NameTextField.value, StageDropdown.value) in NameAndStageList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Такий працівник вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"""UPDATE workers SET Name='{NameTextField.value}', Stage='{StageDropdown.value.replace("'", "''")}', Password='{PasswordTextField.value}' WHERE Id={EditId}""")
                DBConnector.commit()
                CloseDB()

                Page.go('/workers')


#------------------------------------------------------------------------------------------------------------------------


#Machine------------------------------------------------------------------------------------------------------------------------


        def SetStage(Stage, SetStageRoute):
            global ActiveStage
            ActiveStage = Stage
            Page.go(SetStageRoute)

            
        def SetMachineCount(ClickEvent):
            if MachineCountTextField.value == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Кількість машин не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif int(MachineCountTextField.value)>50:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Кількість машин дуже велика!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"""UPDATE machines_info SET Quantity = {MachineCountTextField.value} WHERE Stage = '{ActiveStage}'""")
                DBConnector.commit()
                CloseDB()

                Page.go('/machines')


        def AddProductForMachine():
            OpenDB()
            DBCursor.execute(f"""SELECT Machine FROM products_for_machines WHERE Exist=1""")
            Machines=DBCursor.fetchall()
            CloseDB()
            if MachineDropdown.value == None:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Машину не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif ProductDropdown.value == None:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Продукт не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif (MachineDropdown.value,) in Machines:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("На цій машинці вже є план!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"SELECT Id FROM products WHERE FullName='{ProductDropdown.value}'")
                ProductId=DBCursor.fetchone()[0]

                DBCursor.execute(f"SELECT MAX(Id) FROM products_for_machines")
                Id=DBCursor.fetchone()[0]
                Id=Id+1 if Id != None else 0
                                 
                DBCursor.execute(f"""INSERT INTO products_for_machines VALUES ({Id}, {MachineDropdown.value}, {ProductId}, '{datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}', '?', 1)""")
                DBConnector.commit()
                CloseDB()

                Page.go('/products_for_machines')

        
        def EditProductForMachine():
            OpenDB()
            DBCursor.execute(f"""SELECT Machine FROM products_for_machines WHERE Exist=1 AND Id!={EditId}""")
            Machines=DBCursor.fetchall()
            CloseDB()
            if MachineDropdown.value == None:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Машину не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif ProductDropdown.value == None:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Продукт не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif (MachineDropdown.value,) in Machines:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("На цій машинці вже є план!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"SELECT Id FROM products WHERE FullName='{ProductDropdown.value}'")
                ProductId=DBCursor.fetchone()[0]

                DBCursor.execute(f"""UPDATE products_for_machines SET Machine={MachineDropdown.value}, ProductId={ProductId} WHERE Id={EditId}""")
                DBConnector.commit()
                CloseDB()

                Page.go('/products_for_machines')


#------------------------------------------------------------------------------------------------------------------------

        
#Product------------------------------------------------------------------------------------------------------------------------

        
        def AddProduct():
            OpenDB()
            DBCursor.execute(f"SELECT * FROM products""")
            ProductsData=DBCursor.fetchall()
            CloseDB()
            ProductArtikelList=list(map(lambda ProductInfo: ProductInfo[1], ProductsData))
            ProductFullNamesList=list(map(lambda ProductInfo: ProductInfo[2], ProductsData))
            ProductShortNamesList=list(map(lambda ProductInfo: ProductInfo[3], ProductsData))

            if ArtikelTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Артикул не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif FullNameTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Повне найменуванння не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif ShortNameTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Коротке найменуванння не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif ProductDropdown.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Зв`язаний продукт не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif StageDropdown.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Етап не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif ArtikelTextField.value in ProductArtikelList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Такий артикул вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            elif FullNameTextField.value in ProductFullNamesList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Таке повне найменуванння вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            elif ShortNameTextField.value in ProductShortNamesList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Таке коротке найменуванння вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"SELECT MAX(Id) FROM products")
                Id=DBCursor.fetchone()[0]
                Id=Id+1 if Id != None else 0
                                 
                DBCursor.execute(f"""INSERT INTO products VALUES ({Id}, '{ArtikelTextField.value}', '{FullNameTextField.value}', '{ShortNameTextField.value}', '{ProductDropdown.value}', '{StageDropdown.value.replace("'", "''")}', True)""")
                DBConnector.commit()
                CloseDB()

                Page.go('/products')
        

        def EditProduct():
            OpenDB()
            DBCursor.execute(f"SELECT * FROM products WHERE Id != {EditId}""")
            ProductsData=DBCursor.fetchall()
            CloseDB()
            ProductArtikelList=list(map(lambda ProductInfo: ProductInfo[1], ProductsData))
            ProductFullNamesList=list(map(lambda ProductInfo: ProductInfo[2], ProductsData))
            ProductShortNamesList=list(map(lambda ProductInfo: ProductInfo[3], ProductsData))

            if ArtikelTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Артикул не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif FullNameTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Повне найменуванння не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif ShortNameTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Коротке найменуванння не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif ProductDropdown.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Зв`язаний продукт не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif StageDropdown.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Зв`язаний продукт не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            elif ArtikelTextField.value in ProductArtikelList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Такий артикул вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            elif FullNameTextField.value in ProductFullNamesList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Таке повне найменуванння вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            elif ShortNameTextField.value in ProductShortNamesList:
                def CloseAlert(ClickEvent):
                    ExistAlert.open = False
                    Page.update()

                ExistAlert = flet.AlertDialog(title=flet.Text("Таке коротке найменуванння вже є!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = ExistAlert
                ExistAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"""UPDATE products SET Artikel = '{ArtikelTextField.value}', FullName = '{FullNameTextField.value}', ShortName = '{ShortNameTextField.value}', Come='{ProductDropdown.value}', Stage='{StageDropdown.value.replace("'", "''")}' WHERE Id = {EditId}""")
                DBConnector.commit()
                CloseDB()

                Page.go('/products')


#------------------------------------------------------------------------------------------------------------------------


#Coming------------------------------------------------------------------------------------------------------------------------


        def AddComingProduct(Product):
            global AddedComingProduct
            AddedComingProduct=Product.replace(" (В'язання)", "").replace(" (ПВХ)", "").replace(" (Оверлок)", "").replace(" (Упаковка)", "")
            Page.go("/add_coming_count")


        def AddComing():
            if ComingCountTextField.value == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Кількість приходу не вказано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif int(ComingCountTextField.value)==0:
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Кількість приходу не може дорівнювати 0!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif StageDropdown.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Етап не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            else:
                OpenDB()

                DBCursor.execute(f"""SELECT Product, Stage FROM comings WHERE TimeEnd='?'""")
                Comings=list(map(lambda Element: (Element[0], Element[1]), DBCursor.fetchall()))

                if (AddedComingProduct, StageDropdown.value.replace("'", "''")) not in Comings:
                    DBCursor.execute(f"SELECT MAX(Id) FROM comings")
                    Id=DBCursor.fetchone()[0]
                    Id=Id+1 if Id != None else 0
                    DBCursor.execute(f"""INSERT INTO comings VALUES ({Id}, '{StageDropdown.value.replace("'", "''")}', '{AddedComingProduct}', {ComingCountTextField.value}, '{str(datetime.datetime.now().strftime("%d.%m.%Y %H:%M"))}', '?')""")
                
                    if int(ComingCountTextField.value)==0:
                        DBCursor.execute(f"""UPDATE comings SET TimeEnd='{str(datetime.datetime.now().strftime("%d.%m.%Y %H:%M"))}' WHERE Id={Id}""")
                else:
                    DBCursor.execute(f"""SELECT Id FROM comings WHERE Product='{AddedComingProduct}' AND Stage='{StageDropdown.value.replace("'", "''")}'""")
                    ComeId=DBCursor.fetchone()[0]
                    DBCursor.execute(f"""SELECT Pair FROM comings WHERE Id={ComeId}""")
                    ComePair=DBCursor.fetchone()[0]+int(ComingCountTextField.value)
                    DBCursor.execute(f"""UPDATE comings SET Pair={ComePair}, TimeEnd='?' WHERE Id={ComeId}""")
                
                    if ComePair==0:
                        DBCursor.execute(f"""UPDATE comings SET TimeEnd='{str(datetime.datetime.now().strftime("%d.%m.%Y %H:%M"))}' WHERE Id={ComeId}""")
                
                DBConnector.commit()
                CloseDB()

                Page.go('/comings')


        def EditComing(ClickEvent):
            if PairTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Кількість товару на прихід не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            else:
                OpenDB()
                ComePair=int(PairTextField.value)
                DBCursor.execute(f"""UPDATE comings SET Pair = {ComePair} WHERE Id = {EditId}""")

                if ComePair==0:
                    DBCursor.execute(f"""UPDATE comings SET TimeEnd='{str(datetime.datetime.now().strftime("%d.%m.%Y %H:%M"))}' WHERE Id={EditId}""")
                else:
                    DBCursor.execute(f"""UPDATE comings SET TimeEnd='?' WHERE Id={EditId}""")
                DBConnector.commit()
                CloseDB()

                Page.go('/comings')


#------------------------------------------------------------------------------------------------------------------------


#Schedule------------------------------------------------------------------------------------------------------------------------


        def AddSchedule(AddScheduleEvent):
            if AddScheduleEvent.files!=None:
                print(AddScheduleEvent.files[0].path)
                ScheduleFile = openpyxl.load_workbook(AddScheduleEvent.files[0].path)
                ScheduleSheet = ScheduleFile.worksheets[0]
                MaxScheduleRow = max((ScheduleElement.row for ScheduleElement in ScheduleSheet['A'] if ScheduleElement.value is not None))+1
                OpenDB()
                DBCursor.execute(f"TRUNCATE TABLE schedule")
                
                Schedule = [(ScheduleRow-1, ScheduleSheet[f'A{ScheduleRow}'].value, ScheduleSheet[f'B{ScheduleRow}'].value, ScheduleSheet[f'C{ScheduleRow}'].value) for ScheduleRow in range(1, MaxScheduleRow)]

                DBCursor.executemany("INSERT INTO schedule VALUES (%s, %s, %s, %s)", Schedule)

                    
                DBConnector.commit()

                CloseDB()

                ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Normal", FilterInput, FilterColumn)

                def CloseAlert(ClickEvent):
                    Alert.open = False
                    Page.update()

                Alert = flet.AlertDialog(title=flet.Text(f"""Розклад ({AddScheduleEvent.files[0].path}) доданий""", selectable=True), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = Alert
                Alert.open = True
                Page.update()


        def UploadSchedule(AddScheduleEvent):
            if AddScheduleEvent!=None:
                ScheduleXlsxFile = openpyxl.Workbook()
                ScheduleXlsxSheet = ScheduleXlsxFile.active
                ScheduleXlsxSheet.title = f"Розклад"
                OpenDB()
                DBCursor.execute(f"SELECT WorkerId, ShiftStart, ShiftEnd FROM schedule")
                Schedules=DBCursor.fetchall()
                CloseDB()

                for ScheduleIndex, (WorkerId, ShiftStart, ShiftEnd) in enumerate(Schedules):
                    ScheduleXlsxSheet[f"A{ScheduleIndex+1}"]=WorkerId
                    ScheduleXlsxSheet[f"B{ScheduleIndex+1}"]=ShiftStart
                    ScheduleXlsxSheet[f"C{ScheduleIndex+1}"]=ShiftEnd
                
                ScheduleXlsxFile.save(f"{AddScheduleEvent.path}/Schedule.xlsx")

                def CloseAlert(ClickEvent):
                    Alert.open = False
                    Page.update()

                Alert = flet.AlertDialog(title=flet.Text(f"""Розклад збережений тут - {AddScheduleEvent.path}/Schedule.xlsx""", selectable=True), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = Alert
                Alert.open = True
                Page.update()


#------------------------------------------------------------------------------------------------------------------------


#Salary------------------------------------------------------------------------------------------------------------------------


        def MonthSalaryDownload(Month, Year):
            FolderPicker.on_result = lambda GetMonthSalaryEvent: GetMonthSalary(GetMonthSalaryEvent, Month, Year)
            FolderPicker.get_directory_path(dialog_title=f"Вибери папку для збереження зарплати за {Month} {Year}")


        def GetMonthSalary(GetMonthSalaryEvent, Month, Year):
            if GetMonthSalaryEvent.path != None:
                MonthNumbersByNames = {'Січень': "01", 'Лютий': "02", 'Березень': "03", 'Квітень': "04", 'Травень': "05", 'Червень': "06", 'Липень': "07", 'Серпень': "08", 'Вересень': "09", 'Жовтень': "10", 'Листопад': "11", 'Грудень': "12"}

                OpenDB()
                DBCursor.execute(f"SELECT ShiftId, WorkerId, ShiftPlan, PairPrice FROM salary WHERE ShiftStart LIKE '%{MonthNumbersByNames[Month]}.{Year}%' Order BY WorkerId")
                Salarys=DBCursor.fetchall()

                DBCursor.execute(f"SELECT Id, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments FROM workers_shifts WHERE ShiftStart LIKE '%{MonthNumbersByNames[Month]}.{Year}%'")
                WorkersShiftsInfo = {}
                for Id, ShiftStart, ShiftEnd, ShiftTime, MachinesAdjustments in DBCursor.fetchall():
                    if ShiftEnd!='?':
                        MachinesAdjustments=eval(MachinesAdjustments)
                        ShiftTimeSeconds = int(ShiftTime.split(' ')[0])*3600+int(ShiftTime.split(' ')[2])*60
                        MachinesAdjustmentsInfo = {Machine: 0 for Machine in MachinesAdjustments[0][0].keys()}
                        for MachinesAdjustmentsIndex, MachinesStates in enumerate(MachinesAdjustments):
                            MachineTimeWorkStart = datetime.datetime.strptime(MachinesStates[1], '%Y-%m-%d %H:%M:%S.%f')
                            MachineTimeWorkEnd = datetime.datetime.strptime(MachinesAdjustments[MachinesAdjustmentsIndex+1][1], '%Y-%m-%d %H:%M:%S.%f') if MachinesAdjustmentsIndex != len(MachinesAdjustments) - 1 else (datetime.datetime.strptime(ShiftEnd, '%d.%m.%Y %H:%M'))
                            WorkTime = (MachineTimeWorkEnd - MachineTimeWorkStart).total_seconds()
                            for Machine, MachinesState in MachinesStates[0].items():
                                if MachinesState=="0":
                                    MachinesAdjustmentsInfo[Machine] = MachinesAdjustmentsInfo[Machine]+WorkTime

                        Hours, Minutes = divmod(max(MachinesAdjustmentsInfo.values()) // 60, 60)
                        ActiveTime = f"{int(Hours)} г. {int(Minutes)} хв."

                        Hours, Minutes = divmod((ShiftTimeSeconds-max(MachinesAdjustmentsInfo.values())) // 60, 60)
                        DeactiveTime = f"{int(Hours)} г. {int(Minutes)} хв."
                    else:
                        ActiveTime = f"?"
                        DeactiveTime = f"?"
                    WorkersShiftsInfo[Id] = (ShiftStart, ShiftEnd, ShiftTime, ActiveTime, DeactiveTime)

                DBCursor.execute("SELECT WorkerId, Pair, AddDate FROM workers_gloves_quantity WHERE Sort=1")
                WorkersShiftsPairsInfo = DBCursor.fetchall()

                DBCursor.execute("SELECT Id, Name, Stage FROM workers")
                WorkersNames={Id:(Name, Stage) for Id, Name, Stage in DBCursor.fetchall()}

                SalaryInfo=[]
                for ShiftId, WorkerId, ShiftPlan, PairPrice in Salarys:
                    if ShiftId in WorkersShiftsInfo:
                        ShiftPair=sum(Pair for ShiftWorkerId, Pair, AddDate in WorkersShiftsPairsInfo if ShiftWorkerId==WorkerId and datetime.datetime.strptime(WorkersShiftsInfo[ShiftId][0], "%d.%m.%Y %H:%M") <= datetime.datetime.strptime(AddDate, "%d.%m.%Y %H:%M") <= (datetime.datetime.strptime(WorkersShiftsInfo[ShiftId][1], "%d.%m.%Y %H:%M")))
                        Salary=int((ShiftPair/ShiftPlan)*PairPrice*ShiftPair) if ShiftPlan!=0 else 0
                        SalaryInfo.append((WorkersNames[WorkerId], ShiftPair, Salary, WorkersShiftsInfo[ShiftId][0], WorkersShiftsInfo[ShiftId][1], WorkersShiftsInfo[ShiftId][2], WorkersShiftsInfo[ShiftId][3], WorkersShiftsInfo[ShiftId][4], datetime.datetime.strptime(WorkersShiftsInfo[ShiftId][0], "%d.%m.%Y %H:%M")))
                
                SalaryInfo = list(sorted(SalaryInfo, key=lambda SalaryElement:SalaryElement[-1], reverse=True))
                SalaryInfo = {SalaryWorker:[(SalaryElement[1], SalaryElement[2], SalaryElement[3], SalaryElement[4], SalaryElement[5], SalaryElement[6], SalaryElement[7]) for SalaryElement in SalaryInfo if SalaryElement[0]==SalaryWorker] for SalaryWorker in set([SalaryElement[0] for SalaryElement in SalaryInfo])}
                SalaryInfo = dict(sorted(SalaryInfo.items(), key=lambda SalaryElement:(SalaryElement[0][1], SalaryElement[0][0])))
                
                CloseDB()

                MonthSalaryXlsxFile = openpyxl.Workbook()
                MonthSalaryXlsxSheet = MonthSalaryXlsxFile.active
                MonthSalaryXlsxSheet.title = f"Зарплати за {Month} {Year}"

                MonthSalaryXlsxSheet.column_dimensions["A"].width = 33
                MonthSalaryXlsxSheet.column_dimensions["B"].width = 33
                MonthSalaryXlsxSheet.column_dimensions["C"].width = 33
                MonthSalaryXlsxSheet.column_dimensions["D"].width = 33
                MonthSalaryXlsxSheet.column_dimensions["E"].width = 33
                MonthSalaryXlsxSheet.column_dimensions["F"].width = 33
                MonthSalaryXlsxSheet.column_dimensions["G"].width = 33

                MonthSalaryXlsxSheet["A1"]=f"Всього за {Month} {Year}:"
                MonthSalaryXlsxSheet[f"A1"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["B1"]=f'{sum(sum(Salary[0] for Salary in SalaryElement) for WorkerName, SalaryElement in SalaryInfo.items())} пар'
                MonthSalaryXlsxSheet[f"B1"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["C1"]=f'{sum(sum(Salary[1] for Salary in SalaryElement) for WorkerName, SalaryElement in SalaryInfo.items())} грн'
                MonthSalaryXlsxSheet[f"C1"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["B3"]=f"Кількість пар"
                MonthSalaryXlsxSheet[f"B3"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["C3"]=f"Зарплата"
                MonthSalaryXlsxSheet[f"C3"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["D3"]=f"Зміна"
                MonthSalaryXlsxSheet[f"D3"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["E3"]=f"Час роботи"
                MonthSalaryXlsxSheet[f"E3"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["F3"]=f"Активний час"
                MonthSalaryXlsxSheet[f"F3"].font = Font(color = "1E3559", size=18, bold=True)

                MonthSalaryXlsxSheet["G3"]=f"Деактивний час"
                MonthSalaryXlsxSheet[f"G3"].font = Font(color = "1E3559", size=18, bold=True)

                LastSalaryRow = 6
                SalaryStage =''

                for WorkerName, SalaryElements in SalaryInfo.items():
                    if WorkerName[1]!=SalaryStage:
                        SalaryStage=WorkerName[1]
                        MonthSalaryXlsxSheet[f"A{LastSalaryRow}"]=f'{SalaryStage}'
                        MonthSalaryXlsxSheet[f"A{LastSalaryRow}"].font = Font(color = "1E3559", size=16, bold=True)
                        LastSalaryRow+=1

                        MonthSalaryXlsxSheet[f"A{LastSalaryRow}"]=f"Всього:"
                        MonthSalaryXlsxSheet[f"A{LastSalaryRow}"].font = Font(color = "1E3559", size=16, bold=True)

                        MonthSalaryXlsxSheet[f"B{LastSalaryRow}"]=f'{sum(SalaryElementsSumWorker[0] for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage)} пар'
                        MonthSalaryXlsxSheet[f"B{LastSalaryRow}"].font = Font(color = "1E3559", size=16, bold=True)

                        MonthSalaryXlsxSheet[f"C{LastSalaryRow}"]=f'{sum(SalaryElementsSumWorker[1] for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage)} грн'
                        MonthSalaryXlsxSheet[f"C{LastSalaryRow}"].font = Font(color = "1E3559", size=16, bold=True)

                        Hours, Minutes = sum(int(SalaryElementsSumWorker[4].split()[0]) for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage and SalaryElementsSumWorker[4]!='?'), sum(int(SalaryElementsSumWorker[4].split()[2]) for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage and SalaryElementsSumWorker[4]!='?')
                        Hours += Minutes // 60
                        Minutes %= 60
                        MonthSalaryXlsxSheet[f"E{LastSalaryRow}"]=f"{Hours} г. {Minutes} хв."
                        MonthSalaryXlsxSheet[f"E{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                        Hours, Minutes = sum(int(SalaryElementsSumWorker[5].split()[0]) for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage and SalaryElementsSumWorker[5]!='?'), sum(int(SalaryElementsSumWorker[4].split()[2]) for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage and SalaryElementsSumWorker[4]!='?')
                        Hours += Minutes // 60
                        Minutes %= 60
                        MonthSalaryXlsxSheet[f"F{LastSalaryRow}"]=f"{Hours} г. {Minutes} хв."
                        MonthSalaryXlsxSheet[f"f{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                        Hours, Minutes = sum(int(SalaryElementsSumWorker[6].split()[0]) for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage and SalaryElementsSumWorker[6]!='?'), sum(int(SalaryElementsSumWorker[4].split()[2]) for WorkerNameSum, SalaryElementsSum in SalaryInfo.items() for SalaryElementsSumWorker in SalaryElementsSum if WorkerNameSum[1]==SalaryStage and SalaryElementsSumWorker[4]!='?')
                        Hours += Minutes // 60
                        Minutes %= 60
                        MonthSalaryXlsxSheet[f"G{LastSalaryRow}"]=f"{Hours} г. {Minutes} хв."
                        MonthSalaryXlsxSheet[f"G{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                        LastSalaryRow+=2

                    MonthSalaryXlsxSheet[f"A{LastSalaryRow}"]=f'{WorkerName[0]} ({WorkerName[1]})'
                    MonthSalaryXlsxSheet[f"A{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                    LastSalaryRow+=1

                    MonthSalaryXlsxSheet[f"A{LastSalaryRow}"]=f"Всього:"
                    MonthSalaryXlsxSheet[f"A{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                    MonthSalaryXlsxSheet[f"B{LastSalaryRow}"]=f'{sum(SalaryElement[0] for SalaryElement in SalaryElements)} пар'
                    MonthSalaryXlsxSheet[f"B{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                    MonthSalaryXlsxSheet[f"C{LastSalaryRow}"]=f'{sum(SalaryElement[1] for SalaryElement in SalaryElements)} грн'
                    MonthSalaryXlsxSheet[f"C{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                    Hours, Minutes = sum(int(SalaryElement[4].split()[0]) for SalaryElement in SalaryElements if SalaryElement[4]!='?'), sum(int(SalaryElement[4].split()[2]) for SalaryElement in SalaryElements if SalaryElement[4]!='?')
                    Hours += Minutes // 60
                    Minutes %= 60
                    MonthSalaryXlsxSheet[f"E{LastSalaryRow}"]=f"{Hours} г. {Minutes} хв."
                    MonthSalaryXlsxSheet[f"E{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                    Hours, Minutes = sum(int(SalaryElement[5].split()[0]) for SalaryElement in SalaryElements if SalaryElement[5]!='?'), sum(int(SalaryElement[5].split()[2]) for SalaryElement in SalaryElements if SalaryElement[5]!='?')
                    Hours += Minutes // 60
                    Minutes %= 60
                    MonthSalaryXlsxSheet[f"F{LastSalaryRow}"]=f"{Hours} г. {Minutes} хв."
                    MonthSalaryXlsxSheet[f"F{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                    Hours, Minutes = sum(int(SalaryElement[6].split()[0]) for SalaryElement in SalaryElements if SalaryElement[6]!='?'), sum(int(SalaryElement[6].split()[2]) for SalaryElement in SalaryElements if SalaryElement[6]!='?')
                    Hours += Minutes // 60
                    Minutes %= 60
                    MonthSalaryXlsxSheet[f"G{LastSalaryRow}"]=f"{Hours} г. {Minutes} хв."
                    MonthSalaryXlsxSheet[f"G{LastSalaryRow}"].font = Font(color = "1E3559", size=14, bold=True)

                    LastSalaryRow+=2
                    for Pair, Salary, Start, End, Time, ActiveTime, DeactiveTime in SalaryElements:

                        MonthSalaryXlsxSheet[f"B{LastSalaryRow}"]=f"{Pair} пар"
                        MonthSalaryXlsxSheet[f"B{LastSalaryRow}"].font = Font(color = "1E3559", size=14)

                        MonthSalaryXlsxSheet[f"C{LastSalaryRow}"]=f"{Salary} грн"
                        MonthSalaryXlsxSheet[f"C{LastSalaryRow}"].font = Font(color = "1E3559", size=14)

                        if datetime.datetime.strptime(Start, "%d.%m.%Y %H:%M").date() == datetime.datetime.strptime(End, "%d.%m.%Y %H:%M").date():
                            ShiftText=f"""{datetime.datetime.strptime(Start, "%d.%m.%Y %H:%M").strftime("%d.%m.%Y")}: {datetime.datetime.strptime(Start, "%d.%m.%Y %H:%M").strftime("%H:%M")}-{datetime.datetime.strptime(End, "%d.%m.%Y %H:%M").strftime("%H:%M")}"""
                        else:
                            ShiftText=f"{Start} - {End}"
                        MonthSalaryXlsxSheet[f"D{LastSalaryRow}"]=ShiftText

                        MonthSalaryXlsxSheet[f"E{LastSalaryRow}"]=Time.replace("година", "г.").replace("годин", "г.").replace("хвилина", "хв.").replace("хвилин", "хв.")
                        MonthSalaryXlsxSheet[f"E{LastSalaryRow}"].font = Font(color = "1E3559", size=14)

                        MonthSalaryXlsxSheet[f"F{LastSalaryRow}"]=ActiveTime
                        MonthSalaryXlsxSheet[f"F{LastSalaryRow}"].font = Font(color = "1E3559", size=14)

                        MonthSalaryXlsxSheet[f"G{LastSalaryRow}"]=DeactiveTime
                        MonthSalaryXlsxSheet[f"G{LastSalaryRow}"].font = Font(color = "1E3559", size=14)

                        LastSalaryRow+=2
                    
                    LastSalaryRow+=2



                MonthSalaryXlsxFile.save(f"""{GetMonthSalaryEvent.path}/Зарплати за {Month} {Year}.xlsx""")

                def CloseAlert(ClickEvent):
                    Alert.open = False
                    Page.update()

                Alert = flet.AlertDialog(title=flet.Text(f"""Файл був збережений тут - {GetMonthSalaryEvent.path}/Зарплати за {Month} {Year}.xlsx""", selectable=True), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = Alert
                Alert.open = True
                Page.update()

        
        def SalarySet():
            OpenDB()
            if SalaryEditMonth==None: 
                DBCursor.execute(f"TRUNCATE TABLE salary_info")
            else:                
                MonthNumbersByNames = {'Січень': "01", 'Лютий': "02", 'Березень': "03", 'Квітень': "04", 'Травень': "05", 'Червень': "06", 'Липень': "07", 'Серпень': "08", 'Вересень': "09", 'Жовтень': "10", 'Листопад': "11", 'Грудень': "12"}
    

            for TableRow in RouteTable.rows:
                WorkerId = TableRow.cells[0].content.value.split('.')[0]
                ShiftPlan = TableRow.cells[1].content.value
                PairPrice = TableRow.cells[2].content.value
                if ShiftPlan!='' and PairPrice!='':
                    if PairPrice.replace('.','',1).isdigit():
                        if SalaryEditMonth!=None:
                            DBCursor.execute(f"UPDATE salary SET ShiftPlan={ShiftPlan}, PairPrice={PairPrice} WHERE WorkerId={WorkerId} AND ShiftStart LIKE '%{MonthNumbersByNames[SalaryEditMonth.split(' ')[0]]}.{SalaryEditMonth.split(' ')[1]}%'")
                        else:
                            DBCursor.execute(f"""INSERT INTO salary_info VALUES ({WorkerId}, {ShiftPlan}, {PairPrice})""")
                    else:
                        def CloseAlert(ClickEvent):
                            Alert.open = False
                            Page.update()

                        Alert = flet.AlertDialog(title=flet.Text("Число вказано не вірно!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                        Page.dialog = Alert
                        Alert.open = True
                        Page.update()
                        return True
                elif SalaryEditMonth!=None:
                     DBCursor.execute(f"DELETE FROM salary WHERE WorkerId={WorkerId} AND ShiftStart LIKE '%{MonthNumbersByNames[SalaryEditMonth.split(' ')[0]]}.{SalaryEditMonth.split(' ')[1]}%'")

            DBConnector.commit()
            CloseDB()
            Page.go("/salary")

        
        def EditSalaryMonthInfo(EditMonth):
            global SalaryEditMonth
            SalaryEditMonth=EditMonth
            Page.go("/salary_set")

        
        def EditSalaryShiftInfo(EditShift):
            global SalaryEditShiftId
            SalaryEditShiftId=EditShift
            Page.go("/edit_shift_salary_info")


        def EditShiftSalary():
            if ShiftPlanTextField.value == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Пар за зміну не вказано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif PairPriceTextField.value == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Ціну за пару не вказано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif not str(PairPriceTextField.value).replace('.','',1).isdigit():                
                def CloseAlert(ClickEvent):
                    Alert.open = False
                    Page.update()

                Alert = flet.AlertDialog(title=flet.Text("Число вказано не вірно!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = Alert
                Alert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"UPDATE salary SET ShiftPlan={ShiftPlanTextField.value}, PairPrice={PairPriceTextField.value} WHERE ShiftId={SalaryEditShiftId}")
                DBConnector.commit()
                CloseDB()

                Page.go('/shifts_salary')


#------------------------------------------------------------------------------------------------------------------------


#Unloadings------------------------------------------------------------------------------------------------------------------------


        def EditUnloading(ClickEvent):
            if PairTextField.value.replace(' ', '') == '':
                def CloseAlert(ClickEvent):
                    NameAlert.open = False
                    Page.update()

                NameAlert = flet.AlertDialog(title=flet.Text("Кількість пар не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = NameAlert
                NameAlert.open = True
                Page.update()
            elif ProductTextField.value == None:
                def CloseAlert(ClickEvent):
                    StageAlert.open = False
                    Page.update()

                StageAlert = flet.AlertDialog(title=flet.Text("Продукт не вибрано!"), actions=[flet.TextButton("Ok", on_click=CloseAlert)])
                Page.dialog = StageAlert
                StageAlert.open = True
                Page.update()
            else:
                OpenDB()
                DBCursor.execute(f"SELECT ProductId, Pair FROM unloadings_info WHERE UnloadingId={ActiveId}") 
                ProductsId={ProductId:Pair for ProductId, Pair in DBCursor.fetchall()}

                DBCursor.execute(f"SELECT Id FROM products WHERE ShortName = '{ProductTextField.value}'")
                ProductId=DBCursor.fetchone()[0]
                Pair=int(PairTextField.value)

                if ProductId in ProductsId:
                    DBCursor.execute(f"DELETE FROM unloadings_info WHERE Id = {EditId}")
                    DBCursor.execute(f"""UPDATE unloadings_info SET Pair = {ProductsId[ProductId]+Pair} WHERE UnloadingId = {ActiveId} AND ProductId = {ProductId}""")
                else:
                    DBCursor.execute(f"""UPDATE unloadings_info SET Pair = {Pair}, ProductId={ProductId} WHERE Id = {EditId}""")
                DBConnector.commit()
                CloseDB()

                Page.go('/unloading_info')


#------------------------------------------------------------------------------------------------------------------------


        Page.views.clear()


        Page.views.append(
            flet.View(
                route='/',
                controls=[
                    flet.Text(value="Адміністратор", size=30, color="#28609f", weight="bold"),
                    flet.Container(content=flet.Column(controls=[
                        flet.ElevatedButton(text="Робітники", on_click=lambda RouteEvent: Page.go('/workers'), width=200, height=50, bgcolor="#dde1e7"), 
                        flet.ElevatedButton(text="Машини", on_click=lambda RouteEvent: Page.go('/machines'), width=200, height=50, bgcolor="#dde1e7"), 
                        flet.ElevatedButton(text="Продукти", on_click=lambda RouteEvent: Page.go('/products'), width=200, height=50, bgcolor="#dde1e7"),
                        flet.ElevatedButton(text="Розклад", on_click=lambda RouteEvent: Page.go('/schedule'), width=200, height=50, bgcolor="#dde1e7"), 
                        flet.ElevatedButton(text="Зарплата", on_click=lambda RouteEvent: Page.go('/salary'), width=200, height=50, bgcolor="#dde1e7"), 
                        flet.ElevatedButton(text="Розвантаження", on_click=lambda RouteEvent: Page.go('/unloadings'), width=200, height=50, bgcolor="#dde1e7")
                    ], spacing=30), margin=flet.margin.only(top=40))
                ],
                horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                scroll=flet.ScrollMode.AUTO
            )
        )
 
        PageRoute=Page.route




#Worker------------------------------------------------------------------------------------------------------------------------


        if PageRoute == "/workers":
            CurrentPageStatus["/workers"]=None

            FolderPicker = flet.FilePicker()

            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            WithDeletedCheckbox = flet.Checkbox(label="З видаленими", value=True, on_change=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "WithDeleted", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker"))
            
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="ПІ", key="Name"),
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Пароль", key="Password")
                                    ],
                                )
            
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("ПІ", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Пароль", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )
            

            Page.views.append(
                flet.View(
                    route='/workers',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value="Робітники",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker")
                                    ),
                                    margin=flet.margin.only(left=40)
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.ADD_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: Page.go('/add_worker')
                                    ),
                                    margin=flet.margin.only(left=20)
                                ),
                                flet.Container(
                                    content=flet.ElevatedButton(text="Отримати звіт", on_click=lambda RouteEvent: GetReportDate(), height=40, bgcolor="#dde1e7"),
                                    margin=flet.margin.only(left=20)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "Filter", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                                WithDeletedCheckbox
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable,
                        FolderPicker
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Id", "Name", "Stage", "Password", "Exist"], ["Name", "Stage", "Password"], "/workers", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/worker_shifts", EditRoute="/edit_worker")


        if PageRoute == "/worker_shifts":
            CurrentPageStatus["/worker_shifts"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            OpenDB()
            DBCursor.execute(f"""SELECT Name, Stage, Exist FROM workers WHERE Id = {ActiveId}""")
            WorkerName, WorkerStage, WorkerExist = DBCursor.fetchone()
            CloseDB()

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Початок зміни", key="ShiftStart"),
                                        flet.dropdown.Option(text="Кінець зміни", key="ShiftEnd"),
                                        flet.dropdown.Option(text="Час роботи", key="ShiftTime")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Початок зміни", size=15, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Кінець зміни", size=15, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Час роботи", size=15, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Активний час", size=15, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Деактивний час", size=15, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Машини", size=15, color="#28609f", weight="bold"),
                    ),
                    flet.DataColumn(
                        flet.Text(""),
                    ),
                    flet.DataColumn(
                        flet.Text(""),
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/worker_shifts',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/workers'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Зміни {WorkerName} ({WorkerStage})",
                                    size=30,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000" if WorkerExist==True else '#d1d1d2'
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info")
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Filter", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Id", "ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], ["ShiftStart", "ShiftEnd", "ShiftTime", "ActiveTime", "DeactiveTime"], "/worker_shifts", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift", ShowRoute="/worker_shift_info")


        if PageRoute == "/edit_worker_shift":
            OpenDB()
            DBCursor.execute(f"SELECT ShiftStart, ShiftEnd FROM workers_shifts WHERE Id = {EditId}")
            ShiftStart, ShiftEnd = DBCursor.fetchone()

            DBCursor.execute(f"""SELECT Name, Stage, Exist FROM workers WHERE Id = {ActiveId}""")
            WorkerName, WorkerStage, WorkerExist = DBCursor.fetchone()
            CloseDB()

            ShiftStartTextField = flet.TextField(label="Початок зміни", hint_text="Напиши початок зміни", width=500, value=ShiftStart)
            ShiftEndTextField = flet.TextField(label="Кінець зміни", hint_text="Напиши кінець зміни", width=500, value=ShiftEnd)
            Page.views.append(
                flet.View(
                    route='/edit_worker_shift',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/worker_shifts'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f"Редагувати зміну\n{WorkerName} ({WorkerStage} {ShiftStart}-{ShiftEnd})", size=12, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=ShiftStartTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=ShiftEndTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', bgcolor="#dde1e7", on_click=EditWorkerShift), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/workers_shifts_machines_adjusts":
            OpenDB()
            DBCursor.execute(f"""SELECT ShiftStart, ShiftEnd FROM workers_shifts WHERE Id = {SecondaryActiveId}""")
            ShiftStart, ShiftEnd = DBCursor.fetchone()
            DBCursor.execute(f"""SELECT Name, Stage, Exist FROM workers WHERE Id = {ActiveId}""")
            WorkerName, WorkerStage, WorkerExist = DBCursor.fetchone()
            CloseDB()

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Машина", size=20, color="#28609f", weight="bold"),
                    ),
                    flet.DataColumn(
                        flet.Text("", size=20, color="#28609f", weight="bold"),
                    ),
                    flet.DataColumn(
                        flet.Text("", size=20, color="#28609f", weight="bold"),
                    ),
                    flet.DataColumn(
                        flet.Text("", size=20, color="#28609f", weight="bold"),
                    ),
                    flet.DataColumn(
                        flet.Text("", size=20, color="#28609f", weight="bold"),
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )
            

            Page.views.append(
                flet.View(
                    route='/workers_shifts_machines_adjusts',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/worker_shifts'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Машини працівника\n{WorkerName} ({WorkerStage} {ShiftStart}-{ShiftEnd})",
                                    size=15,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000" if WorkerExist==True else '#d1d1d2'
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: CreateWorkersShiftsMachinesAdjusts()
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            CreateWorkersShiftsMachinesAdjusts()


        if PageRoute == "/worker_shift_info":
            OpenDB()
            DBCursor.execute(f"""SELECT Name, Stage, Exist FROM workers WHERE Id = {ActiveId}""")
            WorkerName, WorkerStage, WorkerExist = DBCursor.fetchone()
            DBCursor.execute(f"""SELECT ShiftStart, ShiftEnd FROM workers_shifts WHERE Id = {SecondaryActiveId}""")
            ShiftStart, ShiftEnd = DBCursor.fetchone()
            CloseDB()
            Page.views.append(
                flet.View(
                    route='/worker_shift_info',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/worker_shifts'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Зміна {WorkerName} ({WorkerStage} {ShiftStart}-{ShiftEnd})",
                                    size=15,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000" if WorkerExist==True else '#d1d1d2'
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Container(content=flet.Row(
                                controls=[
                                    flet.ElevatedButton(text="Виготовлені рукавички працівника", on_click=lambda RouteEvent: Page.go('/worker_shift_gloves'), width=350, height=60, bgcolor="#dde1e7"),
                                    flet.ElevatedButton(text="Виготовлені рукавички машини", on_click=lambda RouteEvent: Page.go('/worker_shift_machine_gloves'), width=350, height=60, bgcolor="#dde1e7")
                                ],
                                alignment=flet.MainAxisAlignment.CENTER,
                                spacing=40
                            ),
                            margin=flet.margin.only(top=50)
                        )
                    ]
                )
            )


        if PageRoute == "/worker_shift_gloves":
            CurrentPageStatus["/worker_shift_gloves"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            OpenDB()
            DBCursor.execute(f"""SELECT Name, Stage, Exist FROM workers WHERE Id = {ActiveId}""")
            WorkerName, WorkerStage, WorkerExist = DBCursor.fetchone()
            DBCursor.execute(f"""SELECT ShiftStart, ShiftEnd FROM workers_shifts WHERE Id = {SecondaryActiveId}""")
            ShiftStart, ShiftEnd = DBCursor.fetchone()
            CloseDB()

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Машина", key="Machine"),
                                        flet.dropdown.Option(text="Продукт", key="Product"),
                                        flet.dropdown.Option(text="Сорт", key="Sort"),
                                        flet.dropdown.Option(text="Пар", key="Pair"),
                                        flet.dropdown.Option(text="Дата додавання", key="AddDate")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Машина", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Продукт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Сорт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves", SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Дата додавання", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves", SortColumnIndex=4)
                    ),
                    flet.DataColumn(
                        flet.Text("", size=20, color="#28609f", weight="bold")
                    ),
                    flet.DataColumn(
                        flet.Text("", size=20, color="#28609f", weight="bold")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/worker_shift_gloves',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/worker_shift_info'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Виготовлені рукавички працівника\n{WorkerName} ({WorkerStage} {ShiftStart}-{ShiftEnd})",
                                    size=15,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000" if WorkerExist==True else '#d1d1d2'
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves")
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Filter", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(['Id', 'Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/worker_shift_gloves", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_worker_shift_gloves")


        if PageRoute == "/worker_shift_machine_gloves":
            CurrentPageStatus["/worker_shift_machine_gloves"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            OpenDB()
            DBCursor.execute(f"""SELECT Name, Stage, Exist FROM workers WHERE Id = {ActiveId}""")
            WorkerName, WorkerStage, WorkerExist = DBCursor.fetchone()
            DBCursor.execute(f"""SELECT ShiftStart, ShiftEnd FROM workers_shifts WHERE Id = {SecondaryActiveId}""")
            ShiftStart, ShiftEnd = DBCursor.fetchone()
            CloseDB()

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Машина", key="Machine"),
                                        flet.dropdown.Option(text="Сорт", key="Sort"),
                                        flet.dropdown.Option(text="Пар", key="Pair")
                                    ],
                                )

            RouteTable = flet.DataTable(
                            sort_column_index=None,
                            sort_ascending=False,
                            columns=[
                                flet.DataColumn(
                                    flet.Text("Машина", size=20, color="#28609f", weight="bold"),
                                    on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Sort', 'Pair'], "/worker_shift_machine_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=0)
                                ),
                                flet.DataColumn(
                                    flet.Text("Сорт", size=20, color="#28609f", weight="bold"),
                                    on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Sort', 'Pair'], "/worker_shift_machine_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=1)
                                ),
                                flet.DataColumn(
                                    flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                                    on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Sort', 'Pair'], "/worker_shift_machine_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=2)
                                )
                            ],
                            rows=[],
                            data_row_max_height=float("infinity")
                        )

            Page.views.append(
                flet.View(
                    route='/worker_shift_machine_gloves',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/worker_shift_info'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Виготовлені рукавички машин\n{WorkerName} ({WorkerStage} {ShiftStart}-{ShiftEnd})",
                                    size=15,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000" if WorkerExist==True else '#d1d1d2'
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(['Machine', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Sort', 'Pair'], "/worker_shift_machine_gloves", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(['Machine', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Sort', 'Pair'], "/worker_shift_machine_gloves", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(['Machine', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Sort', 'Pair'], "/worker_shift_machine_gloves", "Normal", FilterInput, FilterColumn)

            
        if PageRoute == "/edit_worker_shift_gloves":
            OpenDB()
            DBCursor.execute(f"""SELECT Stage FROM workers WHERE Id = {ActiveId}""")
            WorkerStage = DBCursor.fetchone()[0]

            DBCursor.execute(f"SELECT Machine, ProductId, Sort, Pair, AddDate FROM workers_gloves_quantity WHERE Id = {EditId}")
            Machine, ProductId, Sort, Pair, AddDate = DBCursor.fetchone()

            DBCursor.execute(f"SELECT ShortName FROM products WHERE Id={ProductId}")
            Product=DBCursor.fetchone()[0]

            DBCursor.execute(f"""SELECT ShortName FROM products WHERE Stage='{WorkerStage.replace("'", "''")}' ORDER BY ShortName DESC""")
            Products=[Product[0] for Product in DBCursor.fetchall()]
            CloseDB()

            ProductTextField = flet.Dropdown(label="Коротке найменування продукту", hint_text="Вибери коротке найменування продукту", options=[flet.dropdown.Option(Product) for Product in Products], width=500, text_size=12, value=Product)
            GlovesCountTextField = flet.TextField(label="Кількість рукавичок", hint_text="Напиши кількість рукавичок", value=int(Pair*2), suffix_text="шт.", width=400, input_filter=flet.InputFilter(allow=True, regex_string=r"[0-9]*"))
            Page.views.append(
                flet.View(
                    route='/edit_worker_shift_gloves',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/worker_shift_gloves'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Редагувати додані рукавиці працівника({Machine} машина, {Product}, {Sort} сорт, {AddDate})', size=12, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=GlovesCountTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=ProductTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', bgcolor="#dde1e7", on_click=EditWorkerShiftGloves), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/add_worker":
            NameTextField = flet.TextField(label="ПІ", hint_text="Напиши прізвище та ім'я працівника", width=400)
            StageDropdown = flet.Dropdown(label="Етап", hint_text="Вибери етап", options=[flet.dropdown.Option("В'язання"), flet.dropdown.Option('ПВХ'), flet.dropdown.Option('Оверлок'), flet.dropdown.Option('Упаковка')], width=400)
            PasswordTextField = flet.TextField(label="Пароль", hint_text="Напиши пароль працівника", width=400)
            Page.views.append(
                flet.View(
                    route='/add_worker',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/workers'), bgcolor="#dde1e7", icon_size=16), flet.Text(value='Додати працівника', size=30, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=NameTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=StageDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=PasswordTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Додати', bgcolor="#dde1e7", on_click=lambda RouteEvent:AddWorker()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/edit_worker":
            OpenDB()
            DBCursor.execute(f"SELECT Name, Stage, Password FROM workers WHERE Id = {EditId}")
            Name, Stage, Password = DBCursor.fetchone()
            CloseDB()

            NameTextField = flet.TextField(label="ПІ", hint_text="Напиши прізвище та ім'я працівника", value=Name, width=400)
            StageDropdown = flet.Dropdown(label="Етап", hint_text="Вибери етап", value=Stage, options=[flet.dropdown.Option("В'язання"), flet.dropdown.Option('ПВХ'), flet.dropdown.Option('Оверлок'), flet.dropdown.Option('Упаковка')], width=400)
            PasswordTextField = flet.TextField(label="Пароль", value=Password, hint_text="Напиши пароль працівника", width=400)
            Page.views.append(
                flet.View(
                    route='/edit_worker',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/workers'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Редагувати працівника({Name}, {Stage})', size=30, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=NameTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=StageDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=PasswordTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', bgcolor="#dde1e7", on_click=lambda RouteEvent: EditWorker()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


#------------------------------------------------------------------------------------------------------------------------


#Machine------------------------------------------------------------------------------------------------------------------------
        

        if PageRoute == "/machines":
            Page.views.append(
                flet.View(
                    route='/machines',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/'), bgcolor="#dde1e7", icon_size=16), flet.Text(value='Машини', size=30, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=flet.Row(controls=[flet.ElevatedButton(text="В'язання", width=200, height=40, bgcolor="#dde1e7", on_click=lambda ClickEvent: SetStage("В''язання", "/machines_info")), flet.IconButton(icon=flet.icons.MORE_HORIZ_ROUNDED, on_click=lambda RouteEvent: SetStage("В''язання", "/set_machines_info"), bgcolor="#dde1e7", icon_size=16)], alignment=flet.MainAxisAlignment.CENTER, spacing=20), margin=flet.margin.only(top=40)),
                        flet.Container(content=flet.Row(controls=[flet.ElevatedButton(text="ПВХ", width=200, height=40, bgcolor="#dde1e7", on_click=lambda ClickEvent: SetStage("ПВХ", "/stage_machines_gloves")), flet.IconButton(icon=flet.icons.MORE_HORIZ_ROUNDED, on_click=lambda RouteEvent: SetStage("ПВХ", "/set_machines_info"), bgcolor="#dde1e7", icon_size=16)], alignment=flet.MainAxisAlignment.CENTER, spacing=20), margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.Row(controls=[flet.ElevatedButton(text="Оверлок", width=200, height=40, bgcolor="#dde1e7", on_click=lambda ClickEvent: SetStage("Оверлок", "/stage_machines_gloves")), flet.IconButton(icon=flet.icons.MORE_HORIZ_ROUNDED, on_click=lambda RouteEvent: SetStage("Оверлок", "/set_machines_info"), bgcolor="#dde1e7", icon_size=16)], alignment=flet.MainAxisAlignment.CENTER, spacing=20), margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.Row(controls=[flet.ElevatedButton(text="Упаковка", width=200, height=40, bgcolor="#dde1e7", on_click=lambda ClickEvent: SetStage("Упаковка", "/stage_machines_gloves")), flet.IconButton(icon=flet.icons.MORE_HORIZ_ROUNDED, on_click=lambda RouteEvent: SetStage("Упаковка", "/set_machines_info"), bgcolor="#dde1e7", icon_size=16)], alignment=flet.MainAxisAlignment.CENTER, spacing=20), margin=flet.margin.only(top=20)),
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/set_machines_info":
            OpenDB()
            DBCursor.execute(f"""SELECT Quantity FROM machines_info WHERE Stage='{ActiveStage}'""")
            NowMachineCount=DBCursor.fetchone()[0]
            CloseDB()

            MachineCountTextField = flet.TextField(value=str(NowMachineCount) if NowMachineCount != None else '', hint_text="Кількість машин", suffix_text="шт.", input_filter=flet.InputFilter('^[0-9]*'), width=600)
            Page.views.append(
                flet.View(
                    route='/machine_set',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda ChangeRouteEvent: Page.go('/machines'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f"""Кількість машин ({ActiveStage.replace("''", "'")})""", size=30, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        MachineCountTextField,
                        flet.ElevatedButton(text="Запам'ятати", on_click=SetMachineCount, width=200, height=50, bgcolor="#dde1e7")
                    
                    ],
                    spacing=70,
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute=="/machines_info":
            Page.views.append(
                flet.View(
                    route='/machines_info',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/machines'), bgcolor="#dde1e7", icon_size=16), flet.Text(value="Машини в'язання", size=30, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=flet.Row(
                                controls=[
                                    flet.ElevatedButton(text="Виготовлені рукавички", on_click=lambda RouteEvent: Page.go('/stage_machines_gloves'), width=350, height=60, bgcolor="#dde1e7"),
                                    flet.ElevatedButton(text="Продукти до машин", on_click=lambda RouteEvent: Page.go('/products_for_machines'), width=350, height=60, bgcolor="#dde1e7")
                                ],
                                alignment=flet.MainAxisAlignment.CENTER,
                                spacing=40
                            ),
                            margin=flet.margin.only(top=50)
                        )
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/products_for_machines":
            CurrentPageStatus["/products_for_machines"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            WithDeletedCheckbox = flet.Checkbox(label="З видаленими", value=True, on_change=lambda RouteEvent: ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "ProductId", "TimeStart", "TimeEnd"], "/products_for_machines", "WithDeleted", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine"))
            
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Машина", key="Machine"),
                                        flet.dropdown.Option(text="Продукт", key="ProductId"),
                                        flet.dropdown.Option(text="Час початку", key="TimeStart"),
                                        flet.dropdown.Option(text="Час закінчення", key="TimeEnd")
                                    ],
                                )
            
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Машина", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "ProductId", "TimeStart", "TimeEnd"], "/products_for_machines", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Продукт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "ProductId", "TimeStart", "TimeEnd"], "/products_for_machines", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Час початку", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "ProductId", "TimeStart", "TimeEnd"], "/products_for_machines", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Час закінчення", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "ProductId", "TimeStart", "TimeEnd"], "/products_for_machines", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine", SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/products_for_machines',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/machines_info'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value="Продукти до машин в'язання",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "Product", "TimeStart", "TimeEnd"], "/products_for_machines", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine")
                                    ),
                                    margin=flet.margin.only(left=40)
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.ADD_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: Page.go('/add_product_for_machine')
                                    ),
                                    margin=flet.margin.only(left=20)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "Product", "TimeStart", "TimeEnd"], "/products_for_machines", "Filter", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                                WithDeletedCheckbox
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )
            
            ChangePageStatus(["Id", "Machine", "Product", "TimeStart", "TimeEnd", "Exist"], ["Machine", "Product", "TimeStart", "TimeEnd"], "/products_for_machines", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Exist="Exist", Edit=True, Delete=True, EditRoute="/edit_product_for_machine")


        if PageRoute == "/add_product_for_machine":
            OpenDB()
            DBCursor.execute(f"""SELECT Quantity FROM machines_info WHERE Stage='В''язання'""")
            Machines=range(1, DBCursor.fetchone()[0]+1)

            DBCursor.execute(f"SELECT FullName FROM products")
            Products=[Product[0] for Product in DBCursor.fetchall()]
            CloseDB()

            MachineDropdown = flet.Dropdown(label="Машина", hint_text="Вибери машину", options=[flet.dropdown.Option(Machine) for Machine in Machines], width=500, text_size=12)
            ProductDropdown = flet.Dropdown(label="Продукт", hint_text="Вибери продукт", options=[flet.dropdown.Option(Product) for Product in Products], width=500, text_size=12)

            Page.views.append(
                flet.View(
                    route='/add_product_for_machine',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda ChangeRouteEvent: Page.go('/products_for_machines'), bgcolor="#dde1e7", icon_size=16), flet.Text(value='Додати продукт до машини', size=15, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=MachineDropdown, margin=flet.margin.only(top=40)),
                        flet.Container(content=ProductDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Додати', width=200, height=45, bgcolor="#dde1e7", on_click=lambda RouteEvent: AddProductForMachine()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/edit_product_for_machine":
            OpenDB()
            DBCursor.execute(f"""SELECT Machine, ProductId FROM products_for_machines WHERE Id={EditId}""")
            Machine, ProductId = DBCursor.fetchone()
            
            DBCursor.execute(f"SELECT FullName FROM products WHERE Id={ProductId} ORDER BY FullName DESC")
            Product=DBCursor.fetchone()[0]

            DBCursor.execute(f"""SELECT Quantity FROM machines_info WHERE Stage='В''язання'""")
            Machines=range(1, DBCursor.fetchone()[0]+1)

            DBCursor.execute(f"SELECT FullName FROM products ORDER BY FullName DESC")
            Products=[Product[0] for Product in DBCursor.fetchall()]
            CloseDB()

            MachineDropdown = flet.Dropdown(label="Машина", value=Machine, hint_text="Вибери машину", options=[flet.dropdown.Option(Machine) for Machine in Machines], width=500, text_size=12)
            ProductDropdown = flet.Dropdown(label="Продукт", value=Product, hint_text="Вибери продукт", options=[flet.dropdown.Option(Product) for Product in Products], width=500, text_size=12)

            Page.views.append(
                flet.View(
                    route='/edit_product_for_machine',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda ChangeRouteEvent: Page.go('/products_for_machines'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Редагувати продукт до машини ({Machine} машина, {Product})', size=15, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=MachineDropdown, margin=flet.margin.only(top=40)),
                        flet.Container(content=ProductDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', width=200, height=45, bgcolor="#dde1e7", on_click=lambda RouteEvent: EditProductForMachine()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/stage_machines_gloves":
            CurrentPageStatus["/stage_machines_gloves"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Машина", key="Machine"),
                                        flet.dropdown.Option(text="Продукт", key="Product"),
                                        flet.dropdown.Option(text="Сорт", key="Sort"),
                                        flet.dropdown.Option(text="Пар", key="Pair"),
                                        flet.dropdown.Option(text="Дата додавання", key="AddDate")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Машина", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Продукт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Сорт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Дата додавання", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Sort", FilterInput, FilterColumn, SortColumnIndex=4)
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/stage_machines_gloves',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/machines_info'if ActiveStage=="В''язання" else "/machines"),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"""Машини {ActiveStage.lower().replace("''", "'")}""",
                                    size=30,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], ['Machine', 'Product', 'Sort', 'Pair', 'AddDate'], "/stage_machines_gloves", "Normal", FilterInput, FilterColumn)


#------------------------------------------------------------------------------------------------------------------------


#Product------------------------------------------------------------------------------------------------------------------------


        if PageRoute == "/products":
            CurrentPageStatus["/products"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            WithDeletedCheckbox = flet.Checkbox(label="З видаленими", value=True, on_change=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "WithDeleted", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product"))
            
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Артикул", key="Artikel"),
                                        flet.dropdown.Option(text="Повне найменування", key="FullName"),
                                        flet.dropdown.Option(text="Коротке найменування", key="ShortName"),
                                        flet.dropdown.Option(text="Походження", key="Come"),
                                        flet.dropdown.Option(text="Етап", key="Stage")
                                    ],
                                )
            
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Артикул", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Повне ім'я", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Коротке ім'я", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Походження", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product", SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product", SortColumnIndex=4)
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/products',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value="Продукти",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product")
                                    ),
                                    margin=flet.margin.only(left=40)
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.ADD_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: Page.go('/add_product')
                                    ),
                                    margin=flet.margin.only(left=20)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Filter", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                                WithDeletedCheckbox
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )
            
            ChangePageStatus(["Id", "Artikel", "FullName", "ShortName", "Come", "Stage", "Exist"], ["Artikel", "FullName", "ShortName", "Come", "Stage"], "/products", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Exist="Exist", Id="Id", Edit=True, Delete=True, ShowRoute="/product_info", EditRoute="/edit_product")


        if PageRoute == "/product_info":
            CurrentPageStatus["/product_info"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            OpenDB()
            DBCursor.execute(f"""SELECT ShortName, Artikel, Exist FROM products WHERE Id = {ActiveId}""")
            ProductName, ProductArtikel, ProductExist= DBCursor.fetchone()
            CloseDB()

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Машина", key="Machine"),
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Сорт", key="Sort"),
                                        flet.dropdown.Option(text="Пар", key="Pair"),
                                        flet.dropdown.Option(text="Дата додавання", key="AddDate")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Машина", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Сорт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Дата додавання", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=4)
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/product_info',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/products'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"{ProductName} ({ProductArtikel})",
                                    size=15,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000" if ProductExist==True else '#d1d1d2'
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Machine", "Stage", "Sort", "Pair", "AddDate"], ["Machine", "Stage", "Sort", "Pair", "AddDate"], "/product_info", "Normal", FilterInput, FilterColumn)


        if PageRoute == "/add_product":
            OpenDB()
            DBCursor.execute(f"""SELECT FullName FROM products""")
            Products=['Немає']+[Product[0] for Product in DBCursor.fetchall()]
            CloseDB()

            ArtikelTextField = flet.TextField(label="Артикул продукту", hint_text="Напиши артикул продукту", width=500)
            FullNameTextField = flet.TextField(label="Повне найменування продукту", hint_text="Напиши повне найменування продукту", width=500)
            ShortNameTextField = flet.TextField(label="Коротке найменування продукту", hint_text="Напиши коротке найменування продукту", width=500)
            ProductDropdown = flet.Dropdown(label="Зв`язаний продукт", hint_text="Вибери продукт", options=[flet.dropdown.Option(Product) for Product in Products], width=500, text_size=12)
            StageDropdown = flet.Dropdown(label="Етап", hint_text="Вибери етап", options=[flet.dropdown.Option("В'язання"), flet.dropdown.Option('ПВХ'), flet.dropdown.Option('Оверлок'), flet.dropdown.Option('Упаковка')], width=500)

            Page.views.append(
                flet.View(
                    route='/add_product',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/products'), bgcolor="#dde1e7", icon_size=16), flet.Text(value='Додати продукт', size=30, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=ArtikelTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=FullNameTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=ShortNameTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=ProductDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=StageDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Додати', bgcolor="#dde1e7", on_click=lambda RouteEvent:AddProduct()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/edit_product":
            OpenDB()
            DBCursor.execute(f"SELECT Artikel, FullName, ShortName, Come, Stage FROM products WHERE Id = {EditId}")
            Artikel, FullName, ShortName, Come, Stage = DBCursor.fetchone()

            DBCursor.execute(f"""SELECT FullName FROM products""")
            Products=['Немає']+[Product[0] for Product in DBCursor.fetchall()]
            CloseDB()

            ArtikelTextField = flet.TextField(label="Артикул продукту", hint_text="Напиши артикул продукту", width=500, value=Artikel)
            FullNameTextField = flet.TextField(label="Повне найменування продукту", hint_text="Напиши повне найменування продукту", width=500, value=FullName)
            ShortNameTextField = flet.TextField(label="Коротке найменування продукту", hint_text="Напиши коротке найменування продукту", width=500, value=ShortName)
            ProductDropdown = flet.Dropdown(label="Зв`язаний продукт", hint_text="Вибери продукт", options=[flet.dropdown.Option(Product) for Product in Products], width=500, text_size=12, value=Come)
            StageDropdown = flet.Dropdown(label="Етап", hint_text="Вибери етап", options=[flet.dropdown.Option("В'язання"), flet.dropdown.Option('ПВХ'), flet.dropdown.Option('Оверлок'), flet.dropdown.Option('Упаковка')], width=500, value=Stage)

            Page.views.append(
                flet.View(
                    route='/edit_product',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/products'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Редагувати продукт({Artikel}, {FullName}, {ShortName})', size=15, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=ArtikelTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=FullNameTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=ShortNameTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=ProductDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=StageDropdown, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', bgcolor="#dde1e7", on_click=lambda RouteEvent:EditProduct()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


#------------------------------------------------------------------------------------------------------------------------


#Comings------------------------------------------------------------------------------------------------------------------------

        if PageRoute == "/comings":
            CurrentPageStatus["/comings"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            WithDeletedCheckbox = flet.Checkbox(label="З видаленими", value=True, on_change=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "WithDeleted", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming"))
            
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Продукт", key="Product"),
                                        flet.dropdown.Option(text="Пар", key="Pair"),
                                        flet.dropdown.Option(text="Час початку", key="TimeStart"),
                                        flet.dropdown.Option(text="Час закінчення", key="TimeEnd")
                                    ],
                                )
            
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Продукт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Час початку", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming", SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Час закінчення", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Sort", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming", SortColumnIndex=4)
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/comings',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value="Приходи",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming")
                                    ),
                                    margin=flet.margin.only(left=40)
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.ADD_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: Page.go('/add_coming')
                                    ),
                                    margin=flet.margin.only(left=20)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Filter", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                                WithDeletedCheckbox
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )
            
            ChangePageStatus(["Id", "Stage", "Product", "Pair", "TimeStart", "TimeEnd"], ["Stage", "Product", "Pair", "TimeStart", "TimeEnd"], "/comings", "Normal", FilterInput, FilterColumn, WithDeletedCheckbox=WithDeletedCheckbox, Id="Id", Edit=True, ShowRoute="/coming_info", EditRoute="/edit_coming")


        if PageRoute == "/coming_info":
            CurrentPageStatus["/coming_info"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            OpenDB()
            DBCursor.execute(f"""SELECT Product, Stage, TimeStart, TimeEnd FROM comings WHERE Id = {ActiveId}""")
            ComingProduct, ComingStage, ComingTimeStart, ComingTimeEnd = DBCursor.fetchone()
            CloseDB()

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Робітник", key="WorkerName"),
                                        flet.dropdown.Option(text="Продукт", key="Product"),
                                        flet.dropdown.Option(text="Пар", key="Pair"),
                                        flet.dropdown.Option(text="Сорт", key="Sort"),
                                        flet.dropdown.Option(text="Дата додавання", key="AddDate")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Робітник", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Продукт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Sort", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=4)
                    ),
                    flet.DataColumn(
                        flet.Text("Дата додавання", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Sort", FilterInput, FilterColumn, SortColumnIndex=5)
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/coming_info',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/comings'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"{ComingProduct} ({ComingStage}, {ComingTimeStart}-{ComingTimeEnd})",
                                    size=15,
                                    color="#28609f",
                                    weight="bold",
                                    bgcolor="000000" if ComingTimeEnd=="?" else '#d1d1d2'
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], ["Stage", "WorkerName", "Product", "Pair", "Sort", "AddDate"], "/coming_info", "Normal", FilterInput, FilterColumn)


        if PageRoute == "/add_coming":
            OpenDB()
            DBCursor.execute(f"""SELECT ShortName, Stage FROM products""")
            Products = list(map(lambda Element: f"{Element[0]} ({Element[1]})", DBCursor.fetchall()))
            CloseDB()

            Page.views.append(
                flet.View(
                    route='/add_coming',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/comings'), bgcolor="#dde1e7", icon_size=16), flet.Text(value='Додати прихід (Вибери товар для приходу)', size=30, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Column(controls=[flet.ElevatedButton(text=Product, bgcolor="#dde1e7", on_click=lambda RouteEvent, Product=Product:AddComingProduct(Product)) for Product in Products], alignment=flet.CrossAxisAlignment.CENTER)
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/add_coming_count":
            StageDropdown = flet.Dropdown(label="Етап", hint_text="Вибери етап", options=[flet.dropdown.Option("В'язання"), flet.dropdown.Option('ПВХ'), flet.dropdown.Option('Оверлок'), flet.dropdown.Option('Упаковка')], width=500)
            ComingCountTextField = flet.TextField(label="Кількість приходу", hint_text="Напиши кількість приходу", suffix_text="пар", input_filter=flet.InputFilter('^[0-9]*'), width=600)

            Page.views.append(
                flet.View(
                    route='/add_coming_count',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda RouteEvent: Page.go('/add_coming'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Додати прихід (Вибери кількість приходу {AddedComingProduct})', size=15, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=StageDropdown, margin=flet.margin.only(top=40)),
                        flet.Container(content=ComingCountTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Додати', bgcolor="#dde1e7", on_click=lambda RouteEvent:AddComing()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


        if PageRoute == "/edit_coming":
            OpenDB()
            DBCursor.execute(f"SELECT Product, Stage, TimeStart, TimeEnd, Pair FROM comings WHERE Id={EditId}") 
            ComingProduct, ComingStage, ComingTimeStart, ComingTimeEnd, ComingPair=DBCursor.fetchone()
            CloseDB()
            PairTextField = flet.TextField(label="Кількість товару на прихід", hint_text="Напиши кількість товару на прихід", value=int(ComingPair), suffix_text="пар", width=500, input_filter=flet.InputFilter(allow=True, regex_string=r"[0-9]*"))

            Page.views.append(
                flet.View(
                    route='/edit_product',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda ChangeRouteEvent: Page.go('/comings'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Редагувати прихід({ComingProduct}, {ComingStage}, {ComingTimeStart}-{ComingTimeEnd})', size=15, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=PairTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', bgcolor="#dde1e7", on_click=EditComing), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


#------------------------------------------------------------------------------------------------------------------------


#Schedule------------------------------------------------------------------------------------------------------------------------


        if PageRoute == "/schedule":
            CurrentPageStatus["/schedule"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            FolderPicker = flet.FilePicker(on_result=AddSchedule)

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="ПІ", key="Name"),
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Початок зміни", key="ShiftStart"),
                                        flet.dropdown.Option(text="Кінець зміни", key="ShiftEnd"),
                                        flet.dropdown.Option(text="Час роботи", key="ShiftTime")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("ПІ", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Початок зміни", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Кінець зміни", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Час роботи", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=4)
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/schedule',
                    controls = [
                        FolderPicker,
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Розклад",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                ),
                                flet.Container(
                                    content=flet.ElevatedButton(text="Подивитись розклад", on_click=lambda RouteEvent: Page.go('/downloaded_schedule'), height=40, bgcolor="#dde1e7"),
                                    margin=flet.margin.only(left=20)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/schedule", "Normal", FilterInput, FilterColumn)


        if PageRoute == "/downloaded_schedule":
            CurrentPageStatus["/downloaded_schedule"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            FilePicker = flet.FilePicker(on_result=AddSchedule)
            FolderPicker = flet.FilePicker(on_result=UploadSchedule)

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="ПІ", key="Name"),
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Початок зміни", key="ShiftStart"),
                                        flet.dropdown.Option(text="Кінець зміни", key="ShiftEnd"),
                                        flet.dropdown.Option(text="Час роботи", key="ShiftTime")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("ПІ", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Початок зміни", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Кінець зміни", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Sort", FilterInput, FilterColumn, SortColumnIndex=3)
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/downloaded_schedule',
                    controls = [
                        FilePicker,
                        FolderPicker,
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/schedule'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Завантажений розклад",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                flet.Container(
                                    content=flet.ElevatedButton(text="Завантажити розклад", on_click=lambda RouteEvent: FilePicker.pick_files(dialog_title="Вибери розклад xlsx", allowed_extensions=['xlsx']), height=40, bgcolor="#dde1e7"),
                                    margin=flet.margin.only(left=20)
                                ),
                                flet.Container(
                                    content=flet.ElevatedButton(text="Вивантажити розклад", on_click=lambda RouteEvent: FolderPicker.get_directory_path(dialog_title="Вибери папку для вивантаження розклад xlsx"), height=40, bgcolor="#dde1e7"),
                                    margin=flet.margin.only(left=20)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime"], "/downloaded_schedule", "Normal", FilterInput, FilterColumn)


        Page.update()

#------------------------------------------------------------------------------------------------------------------------


#Salary------------------------------------------------------------------------------------------------------------------------


        if PageRoute == "/salary":
            SalaryEditMonth = None
            CurrentPageStatus["/salary"]=None

            FolderPicker = flet.FilePicker()

            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Місяць", key="Name")
                                    ]
                                )
            
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Місяць", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Month"], ["Month"], "/salary", "Sort", FilterInput, FilterColumn, Edit=True, EditRoute='/salary_month_info_edit', SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )
            

            Page.views.append(
                flet.View(
                    route='/salary',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value="Зарплата",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Month"], ["Month"], "/salary", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                ),
                                flet.Container(
                                    content=flet.ElevatedButton(text="Зарплатна інформація", on_click=lambda RouteEvent: Page.go("/salary_set"), height=40, bgcolor="#dde1e7"),
                                    margin=flet.margin.only(left=20)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Month"], ["Month"], "/salary", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7")
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable,
                        FolderPicker
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Month"], ["Month"], "/salary", "Normal", FilterInput, FilterColumn)


        if PageRoute == "/salary_set":
            CurrentPageStatus["/salary_set"]=None

            FilterInput = flet.TextField()
            FilterColumn = flet.Dropdown()
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Робітник", size=20, color="#28609f", weight="bold")
                    ),
                    flet.DataColumn(
                        flet.Text("Пар за зміну", size=20, color="#28609f", weight="bold")
                    ),
                    flet.DataColumn(
                        flet.Text("Ціна за пару", size=20, color="#28609f", weight="bold")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )
            

            Page.views.append(
                flet.View(
                    route='/salary_set',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/salary'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Встановлення зарплати" if SalaryEditMonth==None else f"Встановлення зарплати за {SalaryEditMonth}",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Worker", "ShiftPlan", "PairPrice"], [], "/salary_set", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        RouteTable,
                        flet.ElevatedButton(text="Запам'ятати", on_click=lambda RouteEvent: SalarySet(), width=200, height=50, bgcolor="#dde1e7")
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Worker", "ShiftPlan", "PairPrice"], [], "/salary_set", "Normal", FilterInput, FilterColumn)


        if PageRoute == "/shifts_salary":
            CurrentPageStatus["/shifts_salary"]=None

            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Робітник", key="Name"),
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Початок зміни", key="ShiftStart"),
                                        flet.dropdown.Option(text="Кінець зміни", key="ShiftEnd"),
                                        flet.dropdown.Option(text="Час зміни", key="ShiftTime"),
                                        flet.dropdown.Option(text="Зарплата", key="Salary")
                                    ]
                                )
            
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Робітник", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Sort", FilterInput, FilterColumn, SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Sort", FilterInput, FilterColumn, SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Початок зміни", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Sort", FilterInput, FilterColumn, SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("Кінець зміни", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Sort", FilterInput, FilterColumn, SortColumnIndex=3)
                    ),
                    flet.DataColumn(
                        flet.Text("Час зміни", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Sort", FilterInput, FilterColumn, SortColumnIndex=4)
                    ),
                    flet.DataColumn(
                        flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Sort", FilterInput, FilterColumn, SortColumnIndex=5)
                    ),
                    flet.DataColumn(
                        flet.Text("Зарплата", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Sort", FilterInput, FilterColumn, SortColumnIndex=6)
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )
            

            Page.views.append(
                flet.View(
                    route='/shifts_salary',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/salary'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Зарплата по змінам ({ActiveMonth})",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Normal", FilterInput, FilterColumn)
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Filter", FilterInput, FilterColumn) if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7")
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], ["Name", "Stage", "ShiftStart", "ShiftEnd", "ShiftTime", "ShiftPair", "Salary"], "/shifts_salary", "Normal", FilterInput, FilterColumn)


        if PageRoute == "/edit_shift_salary_info":
            OpenDB()
            DBCursor.execute(f"SELECT ShiftPlan, PairPrice FROM salary WHERE ShiftId={SalaryEditShiftId}") 
            ShiftPlan, PairPrice=DBCursor.fetchone()

            DBCursor.execute(f"SELECT WorkerId, ShiftStart, ShiftEnd FROM workers_shifts WHERE Id={SalaryEditShiftId}") 
            WorkerId, ShiftStart, ShiftEnd=DBCursor.fetchone()

            DBCursor.execute(f"SELECT Name, Stage FROM workers WHERE Id={WorkerId}") 
            WorkerName, WorkerStage=DBCursor.fetchone()
            WorkerName = f'{WorkerName}({WorkerStage})'
            CloseDB()
            ShiftPlanTextField = flet.TextField(value=ShiftPlan, suffix_text="пар", label="Пар за зміну", input_filter=flet.InputFilter('^[0-9]*'), width=400)
            PairPriceTextField = flet.TextField(value=PairPrice, suffix_text="грн", label="Ціна за пару", input_filter=flet.InputFilter('^[0-9.]*'), width=400)

            Page.views.append(
                flet.View(
                    route='/edit_product',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda ChangeRouteEvent: Page.go('/shifts_salary'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Редагувати зарплату({WorkerName}, {ShiftStart}-{ShiftEnd})', size=15, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=ShiftPlanTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=PairPriceTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', bgcolor="#dde1e7", on_click=lambda ChangeRouteEvent: EditShiftSalary()), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

#------------------------------------------------------------------------------------------------------------------------


#Unloadings------------------------------------------------------------------------------------------------------------------------


        if PageRoute == "/unloadings":
            CurrentPageStatus["/unloadings"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            
            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Робітник", key="Name"),
                                        flet.dropdown.Option(text="Етап", key="Stage"),
                                        flet.dropdown.Option(text="Час додавання", key="Date")
                                    ],
                                )
            
            
            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Робітник", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Date"], ["Name", "Stage", "Date"], "/unloadings", "Sort", FilterInput, FilterColumn, Id="Id", Delete=True, ShowRoute="/unloading_info", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Етап", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Date"], ["Name", "Stage", "Date"], "/unloadings", "Sort", FilterInput, FilterColumn, Id="Id", Delete=True, ShowRoute="/unloading_info", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text("Час додавання", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Date"], ["Name", "Stage", "Date"], "/unloadings", "Sort", FilterInput, FilterColumn, Id="Id", Delete=True, ShowRoute="/unloading_info", SortColumnIndex=2)
                    ),
                    flet.DataColumn(
                        flet.Text("")
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/unloadings',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value="Розвантаження",
                                    size=30,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Date"], ["Name", "Stage", "Date"], "/unloadings", "Normal", FilterInput, FilterColumn, Id="Id", Delete=True, ShowRoute="/unloading_info")
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Id", "Name", "Stage", "Date"], ["Name", "Stage", "Date"], "/unloadings", "Filter", FilterInput, FilterColumn, Id="Id", Delete=True, ShowRoute="/unloading_info") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7")
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )
            
            ChangePageStatus(["Id", "Name", "Stage", "Date"], ["Name", "Stage", "Date"], "/unloadings", "Normal", FilterInput, FilterColumn, Id="Id", Delete=True, ShowRoute="/unloading_info")


        if PageRoute == "/unloading_info":
            CurrentPageStatus["/unloading_info"]=None
            PageRow = flet.Row(controls=[], alignment=flet.MainAxisAlignment.CENTER, wrap=True)
            OpenDB()
            DBCursor.execute(f"""SELECT Name, Stage, Date FROM unloadings WHERE Id = {ActiveId}""")
            Name, Stage, Date = DBCursor.fetchone()
            CloseDB()

            FilterInput = flet.TextField(label="Пошук", height=50, width=300)
            FilterColumn = flet.Dropdown(
                                    hint_text="Колонка для пошуку",
                                    options=[
                                        flet.dropdown.Option(text="Продукт", key="Product"),
                                        flet.dropdown.Option(text="Пар", key="Pair")
                                    ],
                                )

            RouteTable = flet.DataTable(
                sort_column_index=None,
                sort_ascending=False,
                columns=[
                    flet.DataColumn(
                        flet.Text("Продукт", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Product", "Pair"], ["Product", "Pair"], "/unloading_info", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_unloading", SortColumnIndex=0)
                    ),
                    flet.DataColumn(
                        flet.Text("Пар", size=20, color="#28609f", weight="bold"),
                        on_sort=lambda RouteEvent: ChangePageStatus(["Id", "Product", "Pair"], ["Product", "Pair"], "/unloading_info", "Sort", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_unloading", SortColumnIndex=1)
                    ),
                    flet.DataColumn(
                        flet.Text(""),
                    ),
                    flet.DataColumn(
                        flet.Text(""),
                    )
                ],
                rows=[],
                data_row_max_height=float("infinity")
            )

            Page.views.append(
                flet.View(
                    route='/unloading_info',
                    controls = [
                        flet.Row(
                            controls=[
                                flet.IconButton(
                                    icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED,
                                    on_click=lambda RouteEvent: Page.go('/unloadings'),
                                    bgcolor="#dde1e7",
                                    icon_size=16
                                ),
                                flet.Text(
                                    value=f"Розвантаження {Name} ({Stage}), {Date}",
                                    size=15,
                                    color="#28609f",
                                    weight="bold"
                                ),
                                flet.Container(
                                    content=flet.IconButton(
                                        icon=flet.icons.AUTORENEW_ROUNDED,
                                        icon_size=16,
                                        bgcolor="#dde1e7",
                                        on_click=lambda RouteEvent: ChangePageStatus(["Id", "Product", "Pair"], ["Product", "Pair"], "/unloading_info", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_unloading")
                                    ),
                                    margin=flet.margin.only(left=40)
                                )
                            ],
                            alignment=flet.MainAxisAlignment.CENTER
                        ),
                        flet.Row(
                            controls=[
                                FilterInput,
                                FilterColumn,
                                flet.ElevatedButton(text="Шукати", on_click=lambda RouteEvent: ChangePageStatus(["Id", "Product", "Pair"], ["Product", "Pair"], "/unloading_info", "Filter", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_unloading") if FilterInput.value != "" and FilterColumn.value not in [None, ''] else None, height=40, bgcolor="#dde1e7"),
                            ],
                            alignment=flet.MainAxisAlignment.CENTER,
                            spacing=25
                        ),
                        PageRow,
                        RouteTable
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )

            ChangePageStatus(["Id", "Product", "Pair"], ["Product", "Pair"], "/unloading_info", "Normal", FilterInput, FilterColumn, Id="Id", Edit=True, Delete=True, EditRoute="/edit_unloading")

        if PageRoute == "/edit_unloading":
            OpenDB()
            DBCursor.execute(f"SELECT ProductId, Pair FROM unloadings_info WHERE Id={EditId}") 
            ProductId, Pair=DBCursor.fetchone()

            DBCursor.execute(f"SELECT ShortName FROM products WHERE Id={ProductId}") 
            ShortNameProduct=DBCursor.fetchone()[0]
            
            DBCursor.execute(f"""SELECT Name, Stage, Date FROM unloadings WHERE Id = {ActiveId}""")
            Name, Stage, Date = DBCursor.fetchone()

            DBCursor.execute(f"""SELECT ShortName FROM products WHERE Stage='Упаковка'""")
            Products=[Product[0] for Product in DBCursor.fetchall()]
            CloseDB()

            ProductTextField = flet.Dropdown(label="Коротке найменування продукту", hint_text="Вибери коротке найменування продукту", options=[flet.dropdown.Option(Product) for Product in Products], width=500, text_size=12, value=ShortNameProduct)
            PairTextField = flet.TextField(label="Кількість пар", hint_text="Напиши кількість пар", value=int(Pair), suffix_text="пар", width=500, input_filter=flet.InputFilter(allow=True, regex_string=r"[0-9]*"))

            Page.views.append(
                flet.View(
                    route='/edit_product',
                    controls=[
                        flet.Row(controls=[flet.IconButton(icon=flet.icons.ARROW_BACK_IOS_NEW_ROUNDED, on_click=lambda ChangeRouteEvent: Page.go('/unloading_info'), bgcolor="#dde1e7", icon_size=16), flet.Text(value=f'Редагувати вивантаження\n({Pair}, {ShortNameProduct})\n({Name}, {Stage}, {Date})', size=15, color="#28609f", weight="bold")], alignment=flet.MainAxisAlignment.CENTER),
                        flet.Container(content=ProductTextField, margin=flet.margin.only(top=40)),
                        flet.Container(content=PairTextField, margin=flet.margin.only(top=20)),
                        flet.Container(content=flet.ElevatedButton(text='Редагувати', bgcolor="#dde1e7", on_click=EditUnloading), margin=flet.margin.only(top=20))
                    ],
                    horizontal_alignment = flet.CrossAxisAlignment.CENTER,
                    scroll=flet.ScrollMode.AUTO
                )
            )


#------------------------------------------------------------------------------------------------------------------------


    def ViewPop(ViewRouteDeleteEvent):
        Page.views.pop()
        Page.go(Page.views[-1].route)



    Page.on_route_change = ChangeRoute
    Page.on_view_pop = ViewPop
    Page.go(Page.route)

flet.app(target=AppScreen)


#virtualenv venv

#source venv/bin/activate

# flet pack GloveStockCountAdmin.py --icon "/Users/apple/Downloads/favicon.png" --name "GloveStockCountAdmin" --product-name "GloveStockCountAdmin" --file-version 1.0

# create-dmg \
#   --volname "GloveStockCountAdmin" \
#   --window-pos 200 120 \
#   --window-size 600 300 \
#   --icon-size 100 \
#   --icon "GloveStockCountAdmin.app" 175 120 \
#   --hide-extension "GloveStockCountAdmin.app" \
#   --app-drop-link 425 120 \
#   "GloveStockCountAdmin.dmg" \
#   "dmg/"
